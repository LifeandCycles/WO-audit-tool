"""
WO Approval Audit Formatter v12.7 — Gate refinements & feedback loop.
Changes from v12.5:
  - Destination gate output: total visits / fees applied / under-threshold summary
  - PR/PRLI gate: consumed part with no PR → Warn (yellow) "confirm van stock"
  - Doc Quality: score ≥60 + missing key element → Warn (not Pass), feedback prompt
  - KEY sheet: added scoring feedback instructions section
Paid/Courtesy: flow-based scoring (arrival → work → closure).
Warranty: same + diagnostic proof required (observed failure + tested/checked).
Destination fee gate: labor-based, visit structure checked, same-customer dedup.
"""
import sys, os, re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ══════════════════════════════════════════════════════════════════════════════
# DOCUMENTATION QUALITY ENGINE  v12.2
# Dual rubric: Paid/Courtesy vs Warranty — silent to the approver
# ══════════════════════════════════════════════════════════════════════════════

# ── Text cleaning ─────────────────────────────────────────────────────────────
def dq_clean(text):
    """Strip Salesforce _x000D_ artifacts and normalize newlines."""
    if not text or str(text).strip() in ("nan", "None", "-", ""):
        return ""
    t = re.sub(r'_x000D_', '\n', str(text))
    t = re.sub(r'\r\n|\r', '\n', t)
    return t.strip()

def dq_word_count(text):
    return len(re.findall(r'\b[a-zA-Z]{3,}\b', text))

# ── Job-type exclusions ───────────────────────────────────────────────────────
DQ_INSTALL = re.compile(
    r'\b(new\s+machine\s+install|new\s+install|machine\s+installation|'
    r'option\s+install|mist\s+condenser\s+install|field\s+install|'
    r'rotary\s+install|hrt\s+install|trt\s+install|new\s+machine\b)',
    re.IGNORECASE)
DQ_PM = re.compile(
    r'\b(comprehensive\s+pm|full\s+pm|pm\s+inspection|pm\s+service|'
    r'preventive\s+maintenance|preventative\s+maintenance|annual\s+pm)\b',
    re.IGNORECASE)
DQ_MOVE = re.compile(
    r'\b(internal\s+move|machine\s+move|relevel\s+after|reinstall\s+after\s+move)\b',
    re.IGNORECASE)
DQ_SELF_REPAIR = re.compile(
    r'customer.{0,80}(repaired|fixed|installed|leveled|anchored|resolved|completed|'
    r'performed|did|handled|replaced).{0,40}'
    r'(themselves|themself|on their own|himself|herself|by themselves)',
    re.IGNORECASE)

# ── Gate thresholds ───────────────────────────────────────────────────────────
DQ_BLOCK_THRESHOLD = 55
DQ_WARN_THRESHOLD  = 70

# ── WO statuses eligible for audit ───────────────────────────────────────────
# Only these statuses get gate rows / report output.
# All other WOs stay in lookups for cross-reference (e.g. destination gate).
AUDIT_STATUSES = {"Field Work Complete", "Parts Reviewed"}

# ── Element 1: Arrival / Finding documented ───────────────────────────────────
DQ_ARRIVAL = re.compile(
    r'\b(found|arrived|upon\s+arrival|on\s+arrival|confirmed\s+alarm|'
    r'customer\s+report|customer\s+states?|customer\s+compla|'
    r'observed|verified\s+(alarm|fault|issue|condition|machine|that)|'
    r'discovered|identified|machine\s+had|history\s+show|alarm\s+history|'
    r'inspected\s+and\s+found|pulled\s+error\s+report|reviewed\s+alarm|'
    r'machine\s+was|upon\s+inspection|at\s+time\s+of\s+arrival|'
    r'machine\s+presented|machine\s+generating|machine\s+showing)\b',
    re.IGNORECASE)

# ── Element 2: Work performed ─────────────────────────────────────────────────
DQ_WORK = re.compile(
    r'\b(removed|replaced|installed|adjusted|cleaned|reseated|re-?seated|'
    r'updated|flashed|homed|reprogrammed|re-?programmed|reset|re-?set|'
    r'torqued|tightened|loosened|lubricated|greased|flushed|purged|ordered|'
    r'swapped|reconnected|disconnected|rewired|re-?wired|repaired|fabricated|'
    r'secured|aligned|calibrated|configured|disabled|enabled|confirmed|'
    r'followed|noted|ensured|powered|assisted|performed|leveled|isolated|'
    r'downloaded|uploaded|initialized|re-?initialized|indicated|activated|'
    r'completed|locked|anchored|armed|re-?armed|geolocated|relocated|'
    r'loaded|captured|cleared|routed|pressed|restarted|re-?booted|set|took|'
    r'provided|used|supplied|applied|replaced|tightened|ran|cycled|tested)\b',
    re.IGNORECASE)

# ── Element 3: Closure confirmed ─────────────────────────────────────────────
DQ_CLOSURE = re.compile(
    r'\b(returned\s+(machine\s+)?(to\s+)?(service|production|operation|customer|normal)|'
    r'returned\s+to\s+normal\s+operation|'
    r'machine\s+(is\s+)?(back\s+in|returned\s+to)\s+(service|production|operation)|'
    r'placed\s+(back\s+)?into\s+(service|production|operation)|'
    r'back\s+in\s+(service|production|operation)|'
    r'no\s+(alarms?|faults?|errors?|issues?|leaks?)\s*(present|found|returned|detected|observed)?|'
    r'confirmed\s+(machine\s+)?(operating|functional|working|running)\s*(normally|properly|correctly)?|'
    r'verified\s+(proper\s+)?(operation|functionality|function)|'
    r'tested\s+(machine\s+)?(operation|function|successfully|good|ok)|'
    r'tested\s+(and\s+)?(functioning|working|operational|confirmed)|'
    r'machine\s+operating\s+normally|operating\s+within\s+normal|'
    r'functioning\s+(normally|properly|correctly)|'
    r'ran\s+(program|mdi|looping|spindle|cycle|lubrication|multiple|test|'
    r'coolant|auger|vibro|ballbar|\d+)|'
    r'cycled\s+(machine|power\s+\d+|through\s+\d+|\d+\s+times?)|'
    r'executed\s+a?\s*program|customer\s+confirmed|'
    r'no\s+further\s+(alarms?|faults?|issues?|problems?)|'
    r'machine\s+back\s+in\s+service|ready\s+for\s+(service|production)|'
    r'fully\s+operational|returned\s+to\s+service)\b',
    re.IGNORECASE)

# ── Element 4 (Warranty): Failure observed/reproduced by tech ─────────────────
DQ_OBSERVED = re.compile(
    r'\b(confirmed\s+alarm|confirmed\s+(the\s+)?(fault|failure|issue|condition|alarm)|'
    r'reproduced\s+(the\s+)?(fault|alarm|issue|condition|problem)|'
    r'verified\s+(the\s+)?(fault|failure|alarm|issue)|'
    r'witnessed\s+(the\s+)?fault|observed\s+(the\s+)?(fault|failure|condition|alarm)|'
    r'found\s+(alarm|fault|condition|machine|the\s+issue|multiple|no\s+issue)|'
    r'alarm\s+history\s+show|pulled\s+error\s+report|error\s+report|'
    r'reviewed\s+alarm\s+history|captured\s+(alarm|error|data)|'
    r'uploaded.*(hbc|haas|work\s+order|salesforce)|'
    r'checklist\s+(completed|filled|uploaded|submitted|populated)|'
    r'sweep\s+and\s+step|pcool\s+sweep|troubleshooting\s+checklist|'
    r'completed\s+the\s+\w+\s+checklist|ran\s+vibro|vibration\s+test|'
    r'i/o\s+bit|bit\s+state|diagnostic\s+screen)\b',
    re.IGNORECASE)

# ── Element 5 (Warranty): Something checked/tested ────────────────────────────
DQ_TESTED = re.compile(
    r'\b(checked|tested|measured|verified|ohmed|inspected|probed|monitored|'
    r'followed\s+hsg|followed\s+haas|per\s+hsg|per\s+haas\s+service|'
    r'per\s+troubleshooting\s+guide|per\s+service\s+alert|'
    r'completed\s+(the\s+)?\w+\s+checklist|filled\s+out\s+(the\s+)?checklist|'
    r'populated\s+checklist|completed\s+checklist|uploaded\s+checklist|'
    r'checklist\s+completed|warranty\s+(review\s+)?checklist|'
    r'uploaded\s+(to\s+)?(hbc|haas|salesforce)|'
    r'captured\s+(p-?cool\s+sweep|data|error\s+report|sweep)|'
    r'ran\s+vibro|vibration\s+analysis|ballbar|'
    r'voltage\s+at|resistance\s+at|ohm|vdc|vac|psi\s+at|'
    r'bit\s+change|i/o\s+(screen|bit|diagnostics)|'
    r'parameter\s+\d+|setting\s+\d+)\b',
    re.IGNORECASE)

# ── Specificity signals (bonus, same for both rubrics) ────────────────────────
DQ_SIG_MEASURE = re.compile(
    r'\b\d+\.?\d*\s*(psi|rpm|mm|volts?|vdc|vac|amps?|ohms?|hz|ms|ft\b|bar|db|'
    r'degrees?|lbs?|nm|in\b)\b'
    r'|\bparameter\s+\d+'
    r'|\b\d+\s*(vac|vdc)\b'
    r'|\b\d*\.\d+\s*ohms?\b'
    r'|l[123n]\s*[-\u2013]\s*(?:l[123n]|g|n)\s*:?\s*\d+'
    r'|\bdc\s+bu[st]s?\b'
    r'|\b\d+\s*[-\u2013]\s*\d+\s*(vac|vdc|v)\b'
    r'|\bwithin\s+spec(ification)?\b|\bin\s+spec(ification)?\b'
    r'|\.\d{3,4}["\x27]'
    r'|\bball\s*bar\b|\bballbar\b|\bvibro\b'
    r'|\bvibration\s+\w+'
    r'|\bi/?o\s+bit\s+(cycled|changed|went|transitioned|toggled)'
    r'|\b\d+\.?\d*\s*seconds?\b|\b\d+\.?\d*\s*sec\b',
    re.IGNORECASE)
DQ_SIG_ALARM   = re.compile(
    r'\balarm\s+[\d][.\d]*\b|\b[\d][.\d]*\s+alarm\b'
    r'|\b(?!20(?:1[5-9]|2[0-9])\b)\d{3,}[.\d]*\s*[-\u2013]\s*[a-zA-Z]'
    r'|\b(?!20(?:1[5-9]|2[0-9])\b)\d{3,}[.\d]*\s*\([a-zA-Z]',
    re.IGNORECASE)
DQ_SIG_SERIAL  = re.compile(
    r'\bsn\s*[-]?\s*\d{4,}\b|\bserial\s*#?\s*\d{4,}\b|\bs/n\s*[-]?\s*\d{4,}\b',
    re.IGNORECASE)
DQ_SIG_PART    = re.compile(r'\b\d{2,3}-[\w-]{3,}\b|part\s*#?\s*\d{2,3}-[\w-]{3,}', re.IGNORECASE)
DQ_SIG_CYCLE   = re.compile(
    r'\b(\d+\s*(tool\s+changes?|cycles?|times?|minutes?|positions?|tests?))\b'
    r'|\b(one|two|three|four|five|six|seven|eight|nine|ten)\s+'
    r'(consecutive\s+)?(\w+\s+){0,3}(tool\s+changes?|cycles?|times?|minutes?|positions?|tests?)\b',
    re.IGNORECASE)
DQ_SIG_CHECKLIST = re.compile(
    r'\b(checklist|sweep\s+and\s+step|error\s+report|uploaded.*hbc|'
    r'uploaded.*haas|uploaded.*salesforce|hbc\s+work\s+order|'
    r'firmware\s+version|AT\d{3,})\b',
    re.IGNORECASE)
DQ_SIG_HSG     = re.compile(
    r'\b(per\s+hsg|per\s+haas\s+service\s+guide|per\s+troubleshooting\s+guide|'
    r'tg\d{4}|per\s+service\s+alert|per\s+haas\s+procedure|'
    r'halo\s+(update|alert|bulletin|notification))\b',
    re.IGNORECASE)

BONUS_SIGNALS = [DQ_SIG_MEASURE, DQ_SIG_ALARM, DQ_SIG_SERIAL,
                 DQ_SIG_PART, DQ_SIG_CYCLE, DQ_SIG_CHECKLIST, DQ_SIG_HSG]
BONUS_NAMES   = ['Measurement', 'Alarm Code', 'Serial Ref',
                 'Part Number', 'Cycle Count', 'Checklist/Upload', 'HSG Ref']

# ── Problem patterns ─────────────────────────────────────────────────────────
DQ_PRIOR_WO_REF = re.compile(
    r'\b(see\s+(prior|previous|ref|reference|wo|work\s+order)\s*(wo|work\s+order|#)?\s*\d+|'
    r'refer\s+to\s+(prior|previous|wo|work\s+order)\s*\d+|'
    r'reference\s+wo\s*[:#]?\s*\d+|'
    r'all\s+diagnostic\s+steps?\s*(and\s+readings?\s+)?in\s+(prior|previous|wo)|'
    r'see\s+previous\s+work\s+order\s+for)\b',
    re.IGNORECASE)
DQ_HSG_PASTE = re.compile(
    r'\b(press\s+prgrm\s+convers|press\s+p\s+enter|p5\s+enter|'
    r'if\s+problem\s+goes\s+away\s+troubleshoot|'
    r'there\s+is\s+no\s+recovery\s+from\s+this\s+condition|'
    r'for\s+machines\s+with\s+a\s+(pc104|maincon)\s*:)\b',
    re.IGNORECASE)
DQ_UNRESOLVED = re.compile(
    r'\b(will\s+report\s+if|will\s+confirm\s+if|'
    r'if\s+this\s+(does\s+not|doesn.t)\s+(correct|fix|resolve|remedy)|'
    r'if\s+(problem|issue|alarm|concern)\s+(returns?|reoccurs?|persists?|comes?\s+back)|'
    r'awaiting\s+results?|to\s+be\s+determined|tbd|'
    r'pending\s+(repair|result|part|installation)|'
    r'monitor\s+for\s+recurrence|follow.?up\s+(required|needed|scheduled)|'
    r'will\s+need\s+to\s+return|may\s+need\s+(to\s+be\s+)?(replaced?|repaired?))\b',
    re.IGNORECASE)

# ── Letter grade ──────────────────────────────────────────────────────────────
def dq_letter_grade(score):
    if score >= 90: return 'A'
    if score >= 80: return 'B'
    if score >= 70: return 'C'
    if score >= 60: return 'D'
    return 'F'

def dq_bonus_signals(combined):
    hits  = [BONUS_NAMES[i] for i, p in enumerate(BONUS_SIGNALS) if p.search(combined)]
    score = 10 if len(hits) >= 3 else (5 if len(hits) >= 1 else 0)
    return score, hits

# ══════════════════════════════════════════════════════════════════════════════
# PAID / COURTESY SCORING
# ══════════════════════════════════════════════════════════════════════════════
def dq_score_paid(cause, ca):
    combined = (cause + ' ' + ca).lower()
    ca_l     = ca.lower()
    words    = dq_word_count(ca)
    bonus, sigs = dq_bonus_signals(combined)
    has_arrival = bool(DQ_ARRIVAL.search(combined))
    has_work    = bool(DQ_WORK.search(ca_l))
    has_closure = bool(DQ_CLOSURE.search(combined))
    unresolved  = bool(DQ_UNRESOLVED.search(combined))
    short_vague = words < 20 and bonus == 0
    e_arrival = 25 if has_arrival else 0
    e_work    = 35 if has_work    else 0
    e_closure = 30 if has_closure else 0
    total     = e_arrival + e_work + e_closure + bonus
    auto_fails = []
    if unresolved:
        auto_fails.append("WO unresolved — outcome not confirmed")
    if not has_arrival:
        auto_fails.append("No arrival/finding documented")
    if not has_work:
        auto_fails.append("No work performed documented")
    if not has_closure:
        auto_fails.append("No closure/confirmation documented")
    if short_vague:
        auto_fails.append(f"Short/vague CA ({words} words) — insufficient for billing")
    return total, {
        'arrival': e_arrival, 'work': e_work, 'closure': e_closure, 'bonus': bonus,
        'has_arrival': has_arrival, 'has_work': has_work, 'has_closure': has_closure,
        'signals': sigs, 'words': words,
    }, auto_fails, []

# ══════════════════════════════════════════════════════════════════════════════
# WARRANTY SCORING
# ══════════════════════════════════════════════════════════════════════════════
def dq_score_warranty(cause, ca):
    combined = (cause + ' ' + ca).lower()
    ca_l     = ca.lower()
    words    = dq_word_count(ca)
    bonus, sigs = dq_bonus_signals(combined)
    has_observed = bool(DQ_OBSERVED.search(combined))
    has_tested   = bool(DQ_TESTED.search(combined))
    has_work     = bool(DQ_WORK.search(ca_l))
    has_closure  = bool(DQ_CLOSURE.search(combined))
    unresolved   = bool(DQ_UNRESOLVED.search(combined))
    prior_wo_ref = bool(DQ_PRIOR_WO_REF.search(combined))
    hsg_paste    = bool(DQ_HSG_PASTE.search(combined))
    short_vague  = words < 20 and bonus == 0
    e_observed = 20 if has_observed else 0
    e_tested   = 25 if has_tested   else 0
    e_work     = 25 if has_work     else 0
    e_closure  = 20 if has_closure  else 0
    total      = e_observed + e_tested + e_work + e_closure + bonus
    auto_fails = []
    if unresolved:
        auto_fails.append("WO unresolved — outcome not confirmed")
    if not has_observed:
        auto_fails.append("Failure not observed/reproduced by tech")
    if not has_tested:
        auto_fails.append("No diagnostic work documented in CA")
    if not has_work:
        auto_fails.append("No work performed documented")
    if not has_closure:
        auto_fails.append("No closure/confirmation documented")
    if short_vague:
        auto_fails.append(f"Short/vague CA ({words} words) — insufficient for warranty claim")
    if prior_wo_ref:
        auto_fails.append("Prior WO referenced instead of documenting in this CA")
    if hsg_paste and not has_observed:
        auto_fails.append("HSG guide pasted — no original diagnostic observations")
    return total, {
        'observed': e_observed, 'tested': e_tested, 'work': e_work,
        'closure': e_closure, 'bonus': bonus,
        'has_observed': has_observed, 'has_tested': has_tested,
        'has_work': has_work, 'has_closure': has_closure,
        'signals': sigs, 'words': words,
    }, auto_fails, []

# ══════════════════════════════════════════════════════════════════════════════
# MAIN GATE FUNCTION
# ══════════════════════════════════════════════════════════════════════════════
C_DQ_EXCELLENT = "375623"
C_DQ_GOOD      = "C6EFCE"
C_DQ_WARN      = "FFD966"
C_DQ_BLOCK     = "9C0006"
C_DQ_SKIP      = "D0D0D0"

def dq_gate(cause_raw, ca_raw, subtype='', wo_num='', has_parts=False):
    cause    = dq_clean(cause_raw)
    ca       = dq_clean(ca_raw)
    combined = (cause + ' ' + ca).lower()
    if DQ_SELF_REPAIR.search(combined):
        return 'Pass', 'N/A — Customer self-repair', {'skip': True, 'reason': 'Customer self-repair'}
    if DQ_INSTALL.search(combined):
        return 'Pass', 'N/A — Install job', {'skip': True, 'reason': 'Install job'}
    if DQ_PM.search(combined):
        return 'Pass', 'N/A — PM job', {'skip': True, 'reason': 'PM job'}
    if DQ_MOVE.search(combined):
        return 'Pass', 'N/A — Machine move/relevel', {'skip': True, 'reason': 'Move/relevel'}
    if not cause and not ca:
        return 'Fail', 'No Cause or Corrective Action documented', {
            'skip': False, 'score_100': 0, 'grade': 'F',
            'auto_fails': ['Missing documentation'], 'signals': [], 'elements': {}
        }
    is_warranty = str(subtype).strip().lower() == 'warranty'
    if is_warranty:
        score_100, elements, auto_fails, warnings = dq_score_warranty(cause, ca)
    else:
        score_100, elements, auto_fails, warnings = dq_score_paid(cause, ca)
    grade = dq_letter_grade(score_100)
    sigs  = elements.get('signals', [])
    has_part_in_ca = bool(DQ_SIG_PART.search(ca))
    if has_parts and not has_part_in_ca:
        warnings.append("Part number not referenced in CA")
    # ── Separate hard blocks from element misses ──────────────────────────
    # Hard blocks always fail regardless of score (unresolved, short/vague,
    # prior WO ref, HSG paste). Element misses (missing arrival, work,
    # closure, observed, tested) become Warn at score ≥60.
    ELEMENT_MISS_PREFIXES = (
        "No arrival", "No work", "No closure", "No diagnostic",
        "Failure not observed",
    )
    hard_blocks = [f for f in auto_fails if not f.startswith(ELEMENT_MISS_PREFIXES)]
    element_misses = [f for f in auto_fails if f.startswith(ELEMENT_MISS_PREFIXES)]

    if score_100 < DQ_BLOCK_THRESHOLD:
        hard_blocks.append(f"Score {score_100}/100 below threshold {DQ_BLOCK_THRESHOLD}")

    score_dict = {
        'skip': False, 'score_100': score_100, 'grade': grade,
        'auto_fails': auto_fails, 'warnings': warnings,
        'signals': sigs, 'elements': elements, 'is_warranty': is_warranty,
    }
    sigs_str = (', '.join(sigs[:3]) + ('...' if len(sigs) > 3 else '')) if sigs else ''
    warn_str = (' | \u26a0 ' + warnings[0]) if warnings else ''

    # Hard blocks → always Fail
    if hard_blocks:
        all_reasons = hard_blocks + element_misses
        detail = f"Grade {grade} ({score_100}/100) | BLOCKED: {'; '.join(all_reasons[:3])}{warn_str}"
        return 'Fail', detail, score_dict

    # Score < block threshold with only element misses → still Fail
    if score_100 < DQ_BLOCK_THRESHOLD and element_misses:
        detail = f"Grade {grade} ({score_100}/100) | BLOCKED: {'; '.join(element_misses[:2])}{warn_str}"
        return 'Fail', detail, score_dict

    # Score ≥60 but missing key elements → Warn with feedback prompt
    if element_misses and score_100 >= 60:
        miss_str = '; '.join(element_misses[:2])
        detail = (f"Grade {grade} ({score_100}/100) | Missing: {miss_str} — "
                  f"verify corrective action and provide feedback{warn_str}")
        return 'Warn', detail, score_dict

    # Score < warn threshold → Warn
    if score_100 < DQ_WARN_THRESHOLD:
        detail = f"Grade {grade} ({score_100}/100) | Passes \u2014 review recommended{warn_str}"
        return 'Warn', detail, score_dict

    detail = f"Grade {grade} ({score_100}/100){(' | ' + sigs_str) if sigs_str else ''}{warn_str}"
    return 'Pass', detail, score_dict

# ══════════════════════════════════════════════════════════════════════════════
# END DOCUMENTATION QUALITY ENGINE
# ══════════════════════════════════════════════════════════════════════════════


C_HEADER_TOP = "0D1F3C"
C_WO_HDR     = "1F3864"
C_PARTS_EVEN = "EBF5FB"
C_PARTS_ODD  = "FFFFFF"
C_SPACER     = "F2F2F2"
C_SUMMARY    = "D6E4F0"
C_PARTS_HDR  = "2E75B6"
C_GOOD       = "A9D08E"
C_BAD        = "FF6B6B"
C_WARN       = "FFD966"
C_MUTED      = "C0C0C0"
C_GOOD_M     = "D5D5D5"
C_BAD_M      = "C0C0C0"
C_WARN_M     = "CACACA"
C_MUTED_TXT  = "808080"
WHITE = "FFFFFF"
DARK  = "1F1F1F"
GOOD_STAT = {"Open", "Approved", "Ready To Ship"}

def solid(h): return PatternFill("solid", fgColor=h)
def mkfont(bold=False, color=WHITE, size=10): return Font(bold=bold, color=color, size=size, name="Arial")
THIN = Side(style="thin", color="AAAAAA")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_BORDER = Border(top=Side(style="thin", color="888888"), bottom=Side(style="thin", color="888888"))

def write_row(ws, rn, fill, bold, fc, vals, h, border=None, nc=25):
    ws.row_dimensions[rn].height = h
    for col in range(1, nc+1):
        c = ws.cell(row=rn, column=col); c.fill=solid(fill); c.font=mkfont(bold=bold, color=fc); c.alignment=Alignment(vertical="center", wrap_text=True)
        if border: c.border=border
    for col, v in vals.items():
        c = ws.cell(row=rn, column=col, value=v); c.fill=solid(fill); c.font=mkfont(bold=bold, color=fc); c.alignment=Alignment(vertical="center", wrap_text=True)
        if border: c.border=border

def write_data(ws, rn, fill, vals, nc=25):
    ws.row_dimensions[rn].height = 16
    for col in range(1, nc+1):
        c = ws.cell(row=rn, column=col); c.fill=solid(fill); c.font=mkfont(color=DARK); c.alignment=Alignment(vertical="center"); c.border=THIN_BORDER
    for col, v in vals.items():
        c = ws.cell(row=rn, column=col, value=v); c.fill=solid(fill); c.font=mkfont(color=DARK); c.border=THIN_BORDER

def setw(ws, w):
    for l, wi in w.items(): ws.column_dimensions[l].width = wi

def val(row, key):
    v = row.get(key, None)
    if v is None: return None
    s = str(v).strip()
    return None if s in ("nan","None","-","") else s

def sf(v):
    try: return float(v)
    except: return 0.0

def hl(ws, r, c, color, bold=False):
    cell = ws.cell(row=r, column=c); cell.fill=solid(color)
    if bold: cell.font=mkfont(bold=True, color=DARK)

def hl_muted(ws, r, c, bold=False):
    cell = ws.cell(row=r, column=c); cell.fill=solid(C_MUTED)
    cell.font=mkfont(bold=bold, color=C_MUTED_TXT)

def apply_hl(ws, r, col_ro, col_st, col_nt, rc, sc, nt, muted):
    if muted:
        if rc: hl_muted(ws, r, col_ro, bold=True)
        if sc: hl_muted(ws, r, col_st)
        if nt and rc: hl_muted(ws, r, col_nt)
    else:
        if rc: hl(ws, r, col_ro, rc, bold=True)
        if sc: hl(ws, r, col_st, sc)
        if nt and rc: hl(ws, r, col_nt, rc)

def build_wo_status_lookup(data):
    lk = {}
    wo_out = data.get("WO_Output")
    if wo_out is not None and len(wo_out) > 0:
        for _, r in wo_out.iterrows(): lk[str(r.get("WO#", ""))] = str(r.get("WO_Status", ""))
    wo_full = data.get("WO")
    if wo_full is not None and len(wo_full) > 0:
        for _, r in wo_full.iterrows():
            wn = str(r.get("WO#", ""))
            if wn not in lk: lk[wn] = str(r.get("Status", ""))
    return lk

def mute(color):
    if color == C_GOOD: return C_GOOD_M
    if color == C_BAD: return C_BAD_M
    if color == C_WARN: return C_WARN_M
    return color

def get_wo_status(data, wo_num):
    wo_full = data.get("WO", None)
    if wo_full is None or len(wo_full) == 0: return None
    m = wo_full[wo_full["WO#"] == wo_num]
    if len(m) > 0: return str(m.iloc[0].get("Status", ""))
    return None

def is_fwc(data, wo_num):
    s = get_wo_status(data, wo_num)
    return s == "Field Work Complete" if s else True

# === DATA LOADING ===
WO_KEY_MAP = {"WO_Output":"WO#", "Parts_Output":"WO#", "RO":"Work Order: Work Order Number",
              "PARTS_SOLD":"Work Order Number", "WO":"Work Order Number",
              "WOLI":"Work Order Number", "SA":"Work Order Number"}
# Full CA sheet — preferred source, falls back to WO sheet if not present
CA_SHEET_NAME = "Work Order Corrective action"
WO_CA_COLS = {"Cause": "Cause", "Corrective Action": "Corrective Action"}

def load_data(f):
    frames = {}
    for sheet, raw in WO_KEY_MAP.items():
        try: df = pd.read_excel(f, sheet_name=sheet, dtype=str)
        except ValueError: frames[sheet]=pd.DataFrame(); continue
        if raw != "WO#" and raw in df.columns: df=df.rename(columns={raw:"WO#"})
        if "WO#" in df.columns: df["WO#"]=df["WO#"].astype(str).str.strip()
        frames[sheet]=df; print("  %-20s: %6d rows" % (sheet, len(df)))
    # Load full CA sheet if present (no character limit)
    try:
        ca_df = pd.read_excel(f, sheet_name=CA_SHEET_NAME, dtype=str)
        ca_df["WO#"] = ca_df["WorkOrderNumber"].astype(str).str.strip().str.split(".").str[0]
        frames["CA_FULL"] = ca_df
        print("  %-20s: %6d rows (full CA)" % (CA_SHEET_NAME[:20], len(ca_df)))
    except Exception:
        frames["CA_FULL"] = pd.DataFrame()

    # Load Work Order Line Item sheet (SF 18-char ID join) for destination gate
    try:
        woli2 = pd.read_excel(f, sheet_name="Work Order Line Item", dtype=str)
        frames["WOLI2"] = woli2
        print("  %-20s: %6d rows (SF ID join)" % ("Work Order Line Item"[:20], len(woli2)))
    except Exception:
        frames["WOLI2"] = pd.DataFrame()

    # Build SF Work Order ID → WO Number map from WO sheet
    wo_id_map = {}
    wo_df = frames.get("WO", pd.DataFrame())
    for _, r in wo_df.iterrows():
        woid  = str(r.get("Work Order ID", "")).strip()
        wonum = str(r.get("WO#", "")).strip().split(".")[0]
        if woid and wonum:
            wo_id_map[woid[:15]] = wonum
    frames["_wo_id_map"] = wo_id_map

    return frames

def _parse_dt(val):
    """Parse datetime from string, return None on failure."""
    try:
        import dateutil.parser
        return dateutil.parser.parse(str(val))
    except:
        return None

def build_destination_lookups(data):
    """
    Build two lookups used by the destination gate:

    woli_lookup  : WO# -> list of {type, start_dt, date, duration_min,
                                    first_travel, travel_home}
    sa_lookup    : WO# -> list of {sa_num, earliest_dt, date, status, tech}
    wo_customer  : WO# -> customer name  (from WO_Output)

    Primary WOLI source: Work Order Line Item sheet (SF ID join, complete data).
    Fallback: legacy WOLI sheet (for WOs not found in Work Order Line Item).
    """
    woli_df  = data.get("WOLI", pd.DataFrame())
    woli2_df = data.get("WOLI2", pd.DataFrame())
    sa_df    = data.get("SA",   pd.DataFrame())
    wo_df    = data.get("WO_Output", pd.DataFrame())
    wo_id_map = data.get("_wo_id_map", {})

    # Customer lookup
    wo_customer = {}
    for _, r in wo_df.iterrows():
        wn = str(r.get("WO#","")).strip().split(".")[0]
        wo_customer[wn] = str(r.get("Customer","")).strip()

    # ── PRIMARY: Work Order Line Item sheet (SF ID join) ─────────────────
    woli_lookup = {}
    if len(woli2_df) > 0 and wo_id_map:
        for _, r in woli2_df.iterrows():
            woid = str(r.get("WorkOrderId","")).strip()
            wn = wo_id_map.get(woid[:15])
            if not wn:
                continue
            typ = str(r.get("Type__c","")).strip().lower()
            start_dt = _parse_dt(r.get("StartDate",""))
            end_dt   = _parse_dt(r.get("EndDate",""))
            if start_dt and end_dt:
                dur = (end_dt - start_dt).total_seconds() / 60.0
            else:
                dur = 0.0
            date = start_dt.date() if start_dt else None
            # travel_home will be inferred below after sorting
            entry = {"type": typ, "start_dt": start_dt, "date": date,
                     "duration_min": dur, "first_travel": False, "travel_home": False}
            woli_lookup.setdefault(wn, []).append(entry)

        # Infer travel direction from chronological order within each date
        for wn, entries in woli_lookup.items():
            entries.sort(key=lambda e: e["start_dt"] or _parse_dt("2000-01-01"))
            by_date = {}
            for e in entries:
                if e["date"]:
                    by_date.setdefault(e["date"], []).append(e)
            for dt, day_entries in by_date.items():
                # Find last labor index for this day
                last_labor_idx = -1
                for i, e in enumerate(day_entries):
                    if e["type"] == "labor":
                        last_labor_idx = i
                # Travel after last labor = travel home
                if last_labor_idx >= 0:
                    for i, e in enumerate(day_entries):
                        if e["type"] == "travel" and i > last_labor_idx:
                            e["travel_home"] = True
                        elif e["type"] == "travel" and i < last_labor_idx:
                            e["first_travel"] = True

        woli2_wos = set(woli_lookup.keys())
        print("  WOLI (primary)    : %6d WOs from Work Order Line Item (SF ID join)" % len(woli2_wos))
    else:
        woli2_wos = set()

    # ── FALLBACK: legacy WOLI sheet (for WOs not in Work Order Line Item) ─
    legacy_count = 0
    for _, r in woli_df.iterrows():
        wn  = str(r.get("WO#","")).strip().split(".")[0]
        if wn in woli2_wos:
            continue   # already have this WO from the primary source
        typ = str(r.get("Type","")).strip().lower()
        try:    dur = float(r.get("Time Duration (Minutes)", 0) or 0)
        except: dur = 0.0
        first_t  = str(r.get("First Travel Of The Day","")).strip().lower() == "true"
        trav_hom = str(r.get("Travel Home","")).strip().lower() == "true"
        start_dt = _parse_dt(r.get("Start Time",""))
        date     = start_dt.date() if start_dt else None
        entry    = {"type": typ, "start_dt": start_dt, "date": date,
                    "duration_min": dur, "first_travel": first_t, "travel_home": trav_hom}
        woli_lookup.setdefault(wn, []).append(entry)
        legacy_count += 1
    if legacy_count:
        legacy_wos = len(set(woli_lookup.keys()) - woli2_wos)
        print("  WOLI (fallback)   : %6d WOs from legacy WOLI sheet" % legacy_wos)

    # SA lookup
    sa_lookup = {}
    for _, r in sa_df.iterrows():
        wn = str(r.get("WO#","")).strip().split(".")[0]
        earliest_dt = _parse_dt(r.get("Earliest Start Permitted",""))
        date        = earliest_dt.date() if earliest_dt else None
        entry = {
            "sa_num":      str(r.get("Appointment Number","")).strip(),
            "earliest_dt": earliest_dt,
            "date":        date,
            "status":      str(r.get("Status","")).strip(),
            "tech":        str(r.get("Service Tech","")).strip(),
        }
        sa_lookup.setdefault(wn, []).append(entry)

    # ── DESTIN-01 lookup: WO# → billable qty from Work Order Line Item ──
    destin_lookup = {}
    if len(woli2_df) > 0 and "Product_Name__c" in woli2_df.columns and wo_id_map:
        for _, r in woli2_df.iterrows():
            pn = str(r.get("Product_Name__c", "")).strip().upper()
            if pn != "DESTIN-01":
                continue
            woid = str(r.get("WorkOrderId", "")).strip()
            wn = wo_id_map.get(woid[:15])
            if not wn:
                continue
            try:
                qty = float(r.get("Billable_Quantity__c", 0) or 0)
            except:
                qty = 0.0
            destin_lookup[wn] = destin_lookup.get(wn, 0) + qty

    return woli_lookup, sa_lookup, wo_customer, destin_lookup

def build_ca_lookup(data):
    """
    Build WO# -> (cause, ca) lookup.
    Prefers CA_FULL sheet (no character limit) over WO sheet.
    Falls back to WO sheet for any WO not found in CA_FULL.
    """
    lk = {}

    # First pass — WO sheet (truncated but covers all WOs)
    wo = data.get("WO", pd.DataFrame())
    for _, r in wo.iterrows():
        wn    = str(r.get("WO#", "")).strip().split(".")[0]
        cause = str(r.get("Cause", ""))              if pd.notna(r.get("Cause"))              else ""
        ca    = str(r.get("Corrective Action", ""))  if pd.notna(r.get("Corrective Action"))  else ""
        lk[wn] = (cause, ca)

    # Second pass — CA_FULL sheet (preferred, overwrites WO sheet values)
    ca_full = data.get("CA_FULL", pd.DataFrame())
    full_count = 0
    for _, r in ca_full.iterrows():
        wn    = str(r.get("WO#", "")).strip().split(".")[0]
        cause = str(r.get("Cause__c", ""))              if pd.notna(r.get("Cause__c"))              else ""
        ca    = str(r.get("Corrective_Action__c", ""))  if pd.notna(r.get("Corrective_Action__c"))  else ""
        if wn and (cause or ca):
            lk[wn] = (cause, ca)
            full_count += 1

    if full_count:
        print(f"  CA_FULL overrides  : {full_count:6d} WOs using full-length CA text")
    return lk

def get_ca(ca_lk, wo_num):
    """Return (cause, ca) for a WO, cleaned of _x000D_ artifacts."""
    wn = str(wo_num).strip().split(".")[0]
    cause_raw, ca_raw = ca_lk.get(wn, ("", ""))
    return dq_clean(cause_raw), dq_clean(ca_raw)

def enrich_desc(pa, sold):
    if sold is None or len(sold)==0: return pa
    lk={}
    for _,r in sold.iterrows():
        d=r.get("Part Description",None)
        if d and str(d) not in ("nan","None",""):
            lk[(str(r.get("WO#","")),str(r.get("Product Name","")))]=str(d)
    ct=0
    for idx,row in pa.iterrows():
        cur=row.get("PartDescription",None)
        if cur is None or str(cur) in ("nan","None",""):
            d=lk.get((str(row.get("WO#","")),str(row.get("PartNumber",""))))
            if d: pa.at[idx,"PartDescription"]=d; ct+=1
    print("  Descriptions: filled %d from PARTS_SOLD" % ct)
    return pa

def enrich_rma(pa, ro):
    if ro is None or len(ro)==0 or len(pa)==0: pa["Vendor_RMA"]=None; return pa
    rl={}
    for _,r in ro.iterrows():
        rl[(str(r.get("WO#","")),str(r.get("Return Order Number","")),str(r.get("Product Name","")))]=str(r.get("Vendor RMA #",""))
    def g(row):
        w=str(row.get("WO#","")); p=str(row.get("PartNumber","")); rr=str(row.get("RO_Numbers",""))
        if not rr or rr in ("nan","None",""): return None
        ms=[]
        for rn in [x.strip() for x in rr.split(",")]:
            m=rl.get((w,rn,p),"")
            if m and m not in ("nan","None"): ms.append(m)
        return ", ".join(ms) if ms else None
    pa["Vendor_RMA"]=pa.apply(g,axis=1)
    print("  Vendor RMA #: enriched %d" % pa["Vendor_RMA"].notna().sum())
    return pa

def get_ro(ro_df, wo, pn):
    if ro_df is None or len(ro_df)==0: return []
    return ro_df[(ro_df["WO#"]==wo)&(ro_df["Product Name"]==pn)].to_dict("records")

# === AUDIT ENGINE ===
def audit(pr, wst, ros, is_complete=True):
    qr=sf(pr.get("QtyRequested",0)); qc=sf(pr.get("QtyConsumed",0))
    pn=str(pr.get("PartNumber","")); has=str(pr.get("Has_RO",""))=="Yes"; iw=(wst=="Warranty")
    if "CORE CHARGE" in pn.upper(): return (None,None,"")
    fc=(qr>0 and qc>0 and qr==qc); nc_flag=(qr>0 and qc==0); pa=(qr>0 and qc>0 and qc<qr)
    wr=[r for r in ros if val(r,"Reason For Return")=="Warranty"]
    cr=[r for r in ros if val(r,"Reason For Return")=="50% Exchange"]
    nr=[r for r in ros if val(r,"Reason For Return")=="NNU"]
    rc=None; n=""
    if fc and iw:
        if wr: rc=C_GOOD; n="Warranty RO OK"
        else: rc=C_BAD; n="Missing Warranty RO"
        if wr and nr: rc=C_WARN; n="Warranty OK + NNU exists (remove NNU)"
    elif fc and cr and not iw:
        rc=C_GOOD; n="Core RO OK"
        if nr: rc=C_WARN; n="Core OK + NNU exists (remove NNU)"
    elif nc_flag:
        if nr:
            rc=C_GOOD; n="NNU RO OK"
            if wr or cr:
                rc=C_WARN; ex=[]
                if wr: ex.append("Warranty")
                if cr: ex.append("Core")
                n="NNU OK + remove "+"/".join(ex)+" RO"
        elif wr or cr: rc=C_BAD; n="Wrong RO type - needs NNU"
        else: rc=C_BAD; n="No RO - needs NNU"
    elif pa: rc=C_WARN; n="Partial consume - review"
    elif fc and not has and not iw: rc=C_GOOD; n="Consumed OK"
    elif fc and has: rc=C_GOOD; n="Consumed + RO"
    if iw and fc and not has and not cr: rc=C_BAD; n="Warranty - no RO"
    rs=str(pr.get("RO_Statuses","")); sc=None
    if rs and rs not in ("nan","None",""):
        ss=[s.strip() for s in rs.split(",")]; vv=[s for s in ss if s]
        if vv: sc=C_GOOD if all(s in GOOD_STAT for s in vv) else C_BAD
    if not is_complete and rc in (C_BAD, C_WARN): rc=C_MUTED; n=n+" [In Progress]" if n else "In Progress"
    return (rc, sc, n)

# === SHEET 1: WO Approval View ===
NC=25
AH=["WO Number","Customer / Part #","Territory / Description","WO Status / PRLI Status","Record Type / PRLI #","Subtype / PR #","Work Type / D365 SO #","Created Date","WO Tech","SA Sched Date","SA Count","Machine","Serial #","# Parts / Qty Req","Qty Consumed","Qty Not Used","Qty Returned","Part Source","Has RO?","RO Numbers","RO Reasons","RO Statuses","Vendor RMA #","Haas WO Id","Audit Note"]
AW={"A":14,"B":26,"C":28,"D":16,"E":14,"F":14,"G":18,"H":13,"I":20,"J":14,"K":9,"L":16,"M":12,"N":10,"O":10,"P":10,"Q":10,"R":14,"S":8,"T":14,"U":14,"V":14,"W":14,"X":14,"Y":28}
WCM={"WO#":1,"Customer":2,"Territory":3,"WO_Status":4,"WO_RecordType":5,"WO_Subtype":6,"WorkType":7,"CreatedDate":8,"WO_Technician":9,"SA_ScheduledDate":10,"SA_Count":11,"MachineModel":12,"SerialNumber":13,"Parts_Lines":14,"Parts_HasAnyRO":19}
PCM={"WO#":1,"PartNumber":2,"PartDescription":3,"PRLI_Status":4,"PRLI_Number":5,"PR_Number":6,"D365_SO_Number":7,"QtyRequested":14,"QtyConsumed":15,"QtyNotUsed":16,"QtyReturned":17,"PartSource":18,"Has_RO":19,"RO_Numbers":20,"RO_Reasons":21,"RO_Statuses":22,"Vendor_RMA":23}

def build_approval(ws, data):
    wo_all=data["WO_Output"]
    wo=wo_all[wo_all["WO_Status"].isin(AUDIT_STATUSES)] if "WO_Status" in wo_all.columns else wo_all
    pa=data["Parts_Output"]; ro=data["RO"]
    wo_stat_lk = build_wo_status_lookup(data)
    ws.freeze_panes="B2"; ws.row_dimensions[1].height=30
    for c2,h in enumerate(AH,1):
        cl=ws.cell(row=1,column=c2,value=h); cl.fill=solid(C_HEADER_TOP); cl.font=mkfont(bold=True,color=WHITE); cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    setw(ws,AW); cur=2
    for _,wrow in wo.iterrows():
        wn=str(wrow.get("WO#","")); wsub=str(wrow.get("WO_Subtype",""))
        wv={idx:wrow.get(k) for k,idx in WCM.items()}
        write_row(ws,cur,C_WO_HDR,True,WHITE,wv,22,border=HDR_BORDER); cur+=1
        wp=pa[pa["WO#"]==wn] if len(pa) else pd.DataFrame()
        for i,(_,p) in enumerate(wp.iterrows()):
            bg=C_PARTS_EVEN if i%2==0 else C_PARTS_ODD
            pv={idx:val(p,k) for k,idx in PCM.items()}
            write_data(ws,cur,bg,pv)
            pn=val(p,"PartNumber") or ""
            rl=get_ro(ro,wn,pn)
            if rl:
                hd=list(set([val(r2,"Work Order: Haas WO Id") for r2 in rl if val(r2,"Work Order: Haas WO Id")]))
                if hd: ws.cell(cur,24).value=", ".join(hd)
            rc,sc,nt=audit(p.to_dict(),wsub,rl)
            muted = wo_stat_lk.get(wn,"") != "Field Work Complete"
            if nt: ws.cell(cur,25).value=nt
            apply_hl(ws,cur,20,22,25,rc,sc,nt,muted)
            cur+=1
        ws.row_dimensions[cur].height=6
        for c2 in range(1,NC+1): ws.cell(cur,c2).fill=solid(C_SPACER)
        cur+=1

# === SHEET 2: WO Summary ===
SH=["WO Number","Customer","Territory","Status","Record Type","Subtype","Work Type","Urgency","Created Date","Technician","SA Sched Date","SA Count","Machine","Serial #","# Parts","Any RO?","Part Categories"]
SCM={"WO#":1,"Customer":2,"Territory":3,"WO_Status":4,"WO_RecordType":5,"WO_Subtype":6,"WorkType":7,"Urgency":8,"CreatedDate":9,"WO_Technician":10,"SA_ScheduledDate":11,"SA_Count":12,"MachineModel":13,"SerialNumber":14,"Parts_Lines":15,"Parts_HasAnyRO":16,"Parts_Categories":17}
SW={"A":14,"B":28,"C":14,"D":18,"E":14,"F":14,"G":22,"H":14,"I":13,"J":20,"K":18,"L":10,"M":14,"N":14,"O":9,"P":10,"Q":22}

def build_summary(ws, wo_raw):
    wo = wo_raw[wo_raw["WO_Status"].isin(AUDIT_STATUSES)] if "WO_Status" in wo_raw.columns else wo_raw
    ws.freeze_panes="B2"; ws.row_dimensions[1].height=28
    for c2,h in enumerate(SH,1):
        cl=ws.cell(row=1,column=c2,value=h); cl.fill=solid(C_HEADER_TOP); cl.font=mkfont(bold=True,color=WHITE); cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    setw(ws,SW)
    for r,(_,row) in enumerate(wo.iterrows(),2):
        ws.row_dimensions[r].height=18
        for k,idx in SCM.items():
            c2=ws.cell(row=r,column=idx,value=row.get(k)); c2.fill=solid(C_SUMMARY); c2.font=mkfont(color=DARK); c2.alignment=Alignment(vertical="center"); c2.border=THIN_BORDER

# === SHEET 3: Parts Detail ===
PDH=["WO Number","Customer","Territory","WO Status","WO Subtype","Part Source","PRLI #","PR #","Part Number","Part Description","Qty Req","Qty Consumed","Qty Not Used","Qty Returned","PRLI Status","D365 SO #","Has RO?","RO Numbers","RO Reasons","RO Statuses","Vendor RMA #","Haas WO Id","Audit Note"]
PDCM={"WO#":1,"Customer":2,"Territory":3,"WO_Status":4,"WO_Subtype":5,"PartSource":6,"PRLI_Number":7,"PR_Number":8,"PartNumber":9,"PartDescription":10,"QtyRequested":11,"QtyConsumed":12,"QtyNotUsed":13,"QtyReturned":14,"PRLI_Status":15,"D365_SO_Number":16,"Has_RO":17,"RO_Numbers":18,"RO_Reasons":19,"RO_Statuses":20,"Vendor_RMA":21}
PDW={"A":14,"B":26,"C":12,"D":18,"E":14,"F":14,"G":12,"H":12,"I":16,"J":26,"K":9,"L":12,"M":12,"N":12,"O":16,"P":18,"Q":9,"R":18,"S":14,"T":16,"U":16,"V":14,"W":28}

def build_parts(ws, pa, wo, ro, data=None):
    ws.freeze_panes="B2"; ws.row_dimensions[1].height=22
    wo_stat_lk = build_wo_status_lookup(data) if data else {}
    for c2,h in enumerate(PDH,1):
        cl=ws.cell(row=1,column=c2,value=h); cl.fill=solid(C_HEADER_TOP); cl.font=mkfont(bold=True,color=WHITE); cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    setw(ws,PDW)
    wl={}
    if wo is not None and len(wo)>0:
        for _,w in wo.iterrows():
            wl[str(w.get("WO#",""))]={"Customer":w.get("Customer",""),"Territory":w.get("Territory",""),"WO_Status":w.get("WO_Status",""),"WO_Subtype":w.get("WO_Subtype","")}
    for r,(_,row) in enumerate(pa.iterrows(),2):
        ws.row_dimensions[r].height=15
        bg=C_PARTS_EVEN if r%2==0 else C_PARTS_ODD
        wn=str(row.get("WO#","")); wi=wl.get(wn,{}); wst=str(wi.get("WO_Subtype",""))
        rd=row.to_dict(); rd["Customer"]=wi.get("Customer",""); rd["Territory"]=wi.get("Territory",""); rd["WO_Status"]=wi.get("WO_Status",""); rd["WO_Subtype"]=wst
        for k,idx in PDCM.items():
            c2=ws.cell(row=r,column=idx,value=rd.get(k)); c2.fill=solid(bg); c2.font=mkfont(color=DARK); c2.alignment=Alignment(vertical="center"); c2.border=THIN_BORDER
        pn=val(row,"PartNumber") or ""
        rl=get_ro(ro,wn,pn)
        if rl:
            hd=list(set([val(r2,"Work Order: Haas WO Id") for r2 in rl if val(r2,"Work Order: Haas WO Id")]))
            if hd: c2=ws.cell(r,22); c2.value=", ".join(hd); c2.fill=solid(bg); c2.font=mkfont(color=DARK); c2.border=THIN_BORDER
        rc,sc,nt=audit(rd,wst,rl)
        muted = wo_stat_lk.get(wn,"") != "Field Work Complete"
        if nt:
            c2=ws.cell(r,23); c2.value=nt; c2.border=THIN_BORDER
            if muted: c2.fill=solid(C_MUTED); c2.font=mkfont(color=C_MUTED_TXT)
            else: c2.fill=solid(rc or bg); c2.font=mkfont(color=DARK)
        apply_hl(ws,r,18,20,23,rc,sc,nt,muted)

# === SHEET 4: KEY / LEGEND ===
def _krow(ws, r, fill, fc, label, desc):
    ws.row_dimensions[r].height = 32
    c1 = ws.cell(row=r, column=2, value=label)
    c1.fill = solid(fill); c1.font = mkfont(bold=True, color=fc); c1.border = THIN_BORDER
    c1.alignment = Alignment(vertical="center")
    c2 = ws.cell(row=r, column=3, value=desc)
    c2.fill = solid(WHITE); c2.font = mkfont(color=DARK); c2.border = THIN_BORDER
    c2.alignment = Alignment(vertical="center", wrap_text=True)

def build_key(ws, has_haas=False):
    ws.sheet_properties.tabColor = "0D1F3C"
    setw(ws, {"A": 4, "B": 30, "C": 60, "D": 4})
    cur = 1
    # Title
    ws.merge_cells("B1:C1")
    c = ws.cell(row=1, column=2, value="WO Approval Audit - Key & Legend")
    c.fill = solid(C_HEADER_TOP); c.font = mkfont(bold=True, color=WHITE, size=14)
    c.alignment = Alignment(vertical="center")
    ws.cell(row=1, column=3).fill = solid(C_HEADER_TOP)
    ws.row_dimensions[1].height = 36
    cur = 3

    # REPORT LAYOUT
    ws.cell(row=cur, column=2, value="REPORT LAYOUT").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    _krow(ws, cur, C_HEADER_TOP, WHITE, "Global Header (Row 1)", "Column labels for the entire report. Frozen so they stay visible while scrolling.")
    cur += 1
    _krow(ws, cur, C_WO_HDR, WHITE, "WO Banner Row", "One row per Work Order showing WO #, customer, territory, status, subtype, tech, machine, dates, and part count.")
    cur += 1
    _krow(ws, cur, C_PARTS_EVEN, DARK, "Part Detail Rows", "Part lines nested under each WO. Alternating light blue / white. Shows part #, description, quantities, PRLI/PR info, RO data, and audit results.")
    cur += 1
    _krow(ws, cur, C_SPACER, DARK, "Spacer", "Thin grey row separating WO blocks for readability.")
    cur += 2

    # RO NUMBERS COLOUR KEY
    ws.cell(row=cur, column=2, value="RO NUMBERS COLUMN - Audit Colour Key").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=3)
    ws.cell(row=cur, column=2, value="Colour-coded based on whether the correct RO exists for this part scenario.").font = mkfont(color=DARK)
    cur += 1
    _krow(ws, cur, C_GOOD, DARK, "Green - Correct RO", "The right RO type exists for the scenario (Warranty RO on consumed warranty part, Core RO on consumed core part, NNU RO on unconsumed part). No action needed.")
    cur += 1
    _krow(ws, cur, C_BAD, DARK, "Red - Problem / Missing", "Required RO is missing, wrong RO type for this scenario, or NNU RO exists on a consumed part. Needs investigation.")
    cur += 1
    _krow(ws, cur, C_WARN, DARK, "Amber - Review Needed", "Part partially consumed or has conflicting ROs (e.g. correct RO exists but a duplicate also remains). Manual review required.")
    cur += 1
    _krow(ws, cur, C_MUTED, DARK, "Grey (Muted) - In Progress WO", "Same audit rules apply but the WO is not yet Field Work Complete (New, Scheduled, or In Progress). Flags are muted grey because the part/RO situation may still change as work continues.")
    cur += 2

    # RO STATUS COLOUR KEY
    ws.cell(row=cur, column=2, value="RO STATUS COLUMN - Independent Colour Key").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=3)
    ws.cell(row=cur, column=2, value="Colour-coded independently based ONLY on the RO status value, regardless of RO type.").font = mkfont(color=DARK)
    cur += 1
    _krow(ws, cur, C_GOOD, DARK, "Green - Good Status", "RO Status is Open, Approved, or Ready To Ship. The RO is being tracked and processed.")
    cur += 1
    _krow(ws, cur, C_BAD, DARK, "Red - Bad Status", "RO Status is Draft or another non-active status. The RO may not be progressing and needs attention.")
    cur += 2

    # AUDIT RULES
    ws.cell(row=cur, column=2, value="AUDIT RULES").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    rules = [
        ("Rule 1: Warranty Consumed", "Qty Req = Qty Consumed on a Warranty job. A Warranty RO should exist. Green if present, Red if missing."),
        ("Rule 2: Core Consumed", "Qty Req = Qty Consumed with a Core (50% Exchange) RO. Core RO should exist. Green if present, Red if missing."),
        ("Rule 3: Not Consumed (NNU)", "Qty Consumed = 0. An NNU RO should exist to return the unused part. Green if NNU RO exists, Red if missing."),
        ("Rule 4: Wrong RO on Unconsumed", "Part not consumed but Warranty or Core RO exists instead of NNU. Red/Amber - remove the wrong RO type, keep only NNU."),
        ("Rule 5: NNU on Consumed Part", "Part fully consumed but NNU RO exists. Amber - NNU should be deleted since part was used."),
        ("Rule 6: Warranty Missing RO", "Part consumed on Warranty job with no RO at all. Red - every consumed warranty part needs a Warranty RO."),
    ]
    for label, desc in rules:
        ws.row_dimensions[cur].height = 42
        _krow(ws, cur, C_SUMMARY, DARK, label, desc)
        cur += 1
    cur += 1

    # AUDIT NOTE VALUES
    ws.cell(row=cur, column=2, value="AUDIT NOTE VALUES").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    notes = [
        (C_GOOD, "Consumed OK", "Part fully consumed, no RO needed (non-warranty, non-core). All good."),
        (C_GOOD, "Consumed + RO", "Part fully consumed and has an RO. All good."),
        (C_GOOD, "Warranty RO OK", "Warranty part consumed and correct Warranty RO exists."),
        (C_GOOD, "Core RO OK", "Core part consumed and correct Core (50% Exchange) RO exists."),
        (C_GOOD, "NNU RO OK", "Part not consumed and correct NNU RO exists for the return."),
        (C_BAD, "Missing Warranty RO", "Consumed on warranty job but no Warranty RO found."),
        (C_BAD, "No RO - needs NNU", "Part not consumed and no RO exists at all. Create NNU RO."),
        (C_BAD, "Wrong RO type - needs NNU", "Part not consumed but has Warranty/Core RO instead of NNU."),
        (C_BAD, "Warranty - no RO", "Consumed on warranty job with no RO of any kind."),
        (C_WARN, "Warranty OK + NNU exists (remove NNU)", "Warranty RO correct but stale NNU RO also present. Delete the NNU."),
        (C_WARN, "Core OK + NNU exists (remove NNU)", "Core RO correct but stale NNU RO also present. Delete the NNU."),
        (C_WARN, "NNU OK + remove Warranty/Core RO", "NNU correct but old Warranty or Core RO still present. Delete the wrong one."),
        (C_WARN, "Partial consume - review", "Part partially consumed. Review ROs manually."),
    ]
    for fill, label, desc in notes:
        ws.row_dimensions[cur].height = 28
        c1 = ws.cell(row=cur, column=2, value=label)
        c1.fill = solid(fill); c1.font = mkfont(bold=True, color=DARK); c1.border = THIN_BORDER
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=cur, column=3, value=desc)
        c2.fill = solid(WHITE); c2.font = mkfont(color=DARK); c2.border = THIN_BORDER
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        cur += 1
    cur += 1

    # SHEETS OVERVIEW
    ws.cell(row=cur, column=2, value="SHEETS IN THIS WORKBOOK").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    sheets = [
        ("WO Approval View", "Primary audit view. WO banners with part rows underneath. RO Numbers and RO Statuses colour-coded. Audit Note explains each flag."),
        ("WO Summary", "One row per WO with high-level info: customer, territory, status, subtype, tech, machine, part count, RO status."),
        ("Parts Detail", "Flat list of all part lines across all WOs. Same audit colour coding. Useful for filtering and sorting."),
        ("Parts + Haas RMA", "Parts detail merged with Haas RMA status data. No WO headers. Matches by Vendor RMA # first, then falls back to Part # + Haas WO Id. Blue Haas columns on right side."),
        ("Key", "This sheet. Explains layout, colour coding, and audit rules."),
    ]
    for label, desc in sheets:
        _krow(ws, cur, C_HEADER_TOP, WHITE, label, desc)
        cur += 1
    cur += 1

    # ADD TO RO COLOUR
    if has_haas:
        ws.cell(row=cur, column=2, value="PARTS + HAAS RMA SHEET - Special Colours").font = mkfont(bold=True, color=DARK, size=12)
        cur += 1
        _krow(ws, cur, "BDD7EE", DARK, "Light Blue - Add to RO", "Part matched to Haas RMA by Part # + Haas WO Id but has no Vendor RMA link yet. The RMA Order # shown needs to be added to the Return Order in the system.")
        cur += 1
        _krow(ws, cur, "4472C4", WHITE, "Blue Header Columns (Q-AA)", "Data sourced from the Haas RMA Status file: RMA Order, Status, RMA Description, Warranty flag, Warranty Type, Core Type, Days Past Due, Due Date, quantities.")
        cur += 2

    # SCORING FEEDBACK INSTRUCTIONS
    ws.cell(row=cur, column=2, value="SCORING FEEDBACK — How to Request a Review").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=3)
    ws.cell(row=cur, column=2, value=(
        "If a Documentation Quality score seems wrong (Warn or Fail on a well-documented WO), "
        "please submit the following details so we can review and improve the scoring model."
    )).font = mkfont(color=DARK)
    ws.cell(row=cur, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[cur].height = 36
    cur += 1
    fb_items = [
        ("1. WO Number", "The Work Order number flagged (e.g. WO-00142587)."),
        ("2. Gate Flagged", "Which gate result you are disputing (e.g. Documentation Quality — Warn)."),
        ("3. Current Score & Grade", "The score and grade shown in the audit (e.g. 72/100, Grade C)."),
        ("4. What Was Flagged Missing", "The specific element(s) the audit says are missing (e.g. 'closure confirmation')."),
        ("5. Full Corrective Action Text", "Copy the COMPLETE corrective action from the Corrective Action report — not the truncated WO field. This is the most important piece for improving accuracy."),
        ("6. Why You Disagree", "Explain what the tech documented that should satisfy the missing element (e.g. 'The tech wrote ran program and confirmed no alarms which is closure')."),
        ("7. Subtype", "Paid, Courtesy, or Warranty — the scoring rubric differs by type."),
    ]
    for label, desc in fb_items:
        ws.row_dimensions[cur].height = 32
        c1 = ws.cell(row=cur, column=2, value=label)
        c1.fill = solid(C_SUMMARY); c1.font = mkfont(bold=True, color=DARK); c1.border = THIN_BORDER
        c1.alignment = Alignment(vertical="center")
        c2 = ws.cell(row=cur, column=3, value=desc)
        c2.fill = solid(WHITE); c2.font = mkfont(color=DARK); c2.border = THIN_BORDER
        c2.alignment = Alignment(vertical="center", wrap_text=True)
        cur += 1
    cur += 1
    ws.merge_cells(start_row=cur, start_column=2, end_row=cur, end_column=3)
    ws.cell(row=cur, column=2, value=(
        "Send this information to your WO approval admin. Each review helps calibrate the scoring "
        "model — the more detail you provide, the faster accuracy improves."
    )).font = mkfont(color=DARK, size=9)
    ws.cell(row=cur, column=2).alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[cur].height = 28
    cur += 2

    # DATA SOURCES
    ws.cell(row=cur, column=2, value="DATA SOURCES").font = mkfont(bold=True, color=DARK, size=12)
    cur += 1
    sources = [
        ("WO_Output", "Work Order header data: customer, status, subtype, tech, machine, part counts."),
        ("Parts_Output", "Part line items: part #, quantities, PRLI/PR info, RO linkage, return categories."),
        ("RO", "Return Order detail: RO number, status, reason, vendor RMA #, Haas WO Id."),
        ("PARTS_SOLD", "Used to enrich missing part descriptions by matching WO# + Part Number."),
        ("rma_status_data (Haas)", "Haas RMA status export. Matched to parts by Vendor RMA # or Part # + Haas WO Id."),
    ]
    for label, desc in sources:
        _krow(ws, cur, C_SUMMARY, DARK, label, desc)
        cur += 1

# === MAIN ===
# === SHEET 5: Parts + Haas RMA Merged ===
RMA_HEADERS = ["WO Number","Customer","Territory","WO Subtype","Part Number","Part Description",
    "Qty Req","Qty Consumed","Qty Not Used","Has RO?","RO Numbers","RO Reasons","RO Statuses",
    "Vendor RMA #","Haas WO Id","Audit Note",
    "Haas RMA Order","Haas Status","Haas RMA Desc.","Haas Warranty","Haas Warranty Type",
    "Haas Core Type","Haas Days Past Due","Haas Due Date","Haas Open Qty","Haas Pending Qty","Haas Closed Qty"]
RMA_COL_MAP = {"WO#":1,"Customer":2,"Territory":3,"WO_Subtype":4,"PartNumber":5,"PartDescription":6,
    "QtyRequested":7,"QtyConsumed":8,"QtyNotUsed":9,"Has_RO":10,"RO_Numbers":11,"RO_Reasons":12,
    "RO_Statuses":13,"Vendor_RMA":14,"Haas_WO_Id":15,"Audit_Note":16}
RMA_HAAS_COL = {"RMA Order":17,"Status":18,"RMA Desc.":19,"Warranty":20,"Warranty Type":21,
    "Core Type":22,"Days Past Due":23,"Due Date":24,"Open Qty":25,"Pending Qty":26,"Closed Qty":27}
RMA_WIDTHS = {"A":12,"B":24,"C":12,"D":12,"E":16,"F":28,"G":8,"H":10,"I":10,"J":8,"K":14,"L":14,
    "M":14,"N":16,"O":14,"P":30,"Q":14,"R":10,"S":22,"T":10,"U":20,"V":18,"W":12,"X":12,
    "Y":9,"Z":9,"AA":9}
C_HAAS_HDR = "4472C4"
C_HAAS_EVEN = "D6E4F0"
C_HAAS_ODD = "EBF5FB"
C_ADD_RO = "BDD7EE"

def build_rma_merged(ws, data, haas_df):
    pa = data["Parts_Output"]; wo_df = data["WO_Output"]; ro_df = data["RO"]
    wo_stat_lk = build_wo_status_lookup(data)
    ws.freeze_panes = "B2"
    ws.row_dimensions[1].height = 30
    nc = len(RMA_HEADERS)
    # Headers - parts columns use standard header, Haas columns use blue
    for i, h in enumerate(RMA_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        if i <= 16:
            c.fill = solid(C_HEADER_TOP)
        else:
            c.fill = solid(C_HAAS_HDR)
        c.font = mkfont(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    setw(ws, RMA_WIDTHS)

    # Build lookups
    wl = {}
    if wo_df is not None and len(wo_df) > 0:
        for _, w in wo_df.iterrows():
            wl[str(w.get("WO#", ""))] = {"Customer": w.get("Customer", ""), "Territory": w.get("Territory", ""), "WO_Subtype": w.get("WO_Subtype", "")}

    haas_by_rma = {}
    if haas_df is not None:
        for _, r in haas_df.iterrows():
            rma = str(r.get("RMA Order", "")).strip()
            if rma and rma != "nan":
                haas_by_rma.setdefault(rma, []).append(r)

    haas_by_part_wo = {}
    if haas_df is not None:
        for _, r in haas_df.iterrows():
            mat = str(r.get("Material", "")).strip()
            woid = str(r.get("WorkOrder", "")).strip()
            if mat and mat != "nan" and woid and woid != "nan":
                haas_by_part_wo.setdefault((mat, woid), []).append(r)

    ro_haas_ids = {}
    if ro_df is not None and len(ro_df) > 0:
        for _, r in ro_df.iterrows():
            wnum = str(r.get("WO#", "")).strip()
            pn = str(r.get("Product Name", "")).strip()
            hid = str(r.get("Work Order: Haas WO Id", "")).strip()
            if hid and hid != "nan":
                ro_haas_ids.setdefault((wnum, pn), set()).add(hid)
                ro_haas_ids.setdefault((wnum, "__ALL__"), set()).add(hid)

    ro_vrma_lookup = {}
    if ro_df is not None and len(ro_df) > 0:
        for _, r in ro_df.iterrows():
            wnum = str(r.get("WO#", "")).strip()
            ro_num = str(r.get("Return Order Number", "")).strip()
            pn = str(r.get("Product Name", "")).strip()
            vrma = str(r.get("Vendor RMA #", "")).strip()
            if vrma and vrma != "nan":
                ro_vrma_lookup[(wnum, ro_num, pn)] = vrma

    cur = 2
    for _, row in pa.iterrows():
        pn_str = str(row.get("PartNumber", ""))
        if "CORE CHARGE" in pn_str.upper():
            continue
        wn = str(row.get("WO#", ""))
        wi = wl.get(wn, {})
        wst = str(wi.get("WO_Subtype", ""))
        bg = C_HAAS_EVEN if cur % 2 == 0 else C_HAAS_ODD

        rd = row.to_dict()
        rd["Customer"] = wi.get("Customer", "")
        rd["Territory"] = wi.get("Territory", "")
        rd["WO_Subtype"] = wst

        # Get Haas WO Id
        hids = ro_haas_ids.get((wn, pn_str), set())
        if not hids:
            hids = ro_haas_ids.get((wn, "__ALL__"), set())
        rd["Haas_WO_Id"] = ", ".join(hids) if hids else None

        # Audit
        ro_lines = get_ro(ro_df, wn, pn_str)
        rc, sc, nt = audit(rd, wst, ro_lines)
        rd["Audit_Note"] = nt

        # Write part columns
        for col in range(1, nc + 1):
            c = ws.cell(row=cur, column=col)
            c.fill = solid(bg); c.font = mkfont(color=DARK); c.alignment = Alignment(vertical="center"); c.border = THIN_BORDER
        for k, idx in RMA_COL_MAP.items():
            c = ws.cell(row=cur, column=idx, value=rd.get(k))
            c.fill = solid(bg); c.font = mkfont(color=DARK); c.border = THIN_BORDER

        # Audit colours
        muted = wo_stat_lk.get(wn, "") != "Field Work Complete"
        apply_hl(ws, cur, 11, 13, 16, rc, sc, nt, muted)


        # Match Haas RMA - primary: Vendor RMA
        haas_matched = None
        add_to_ro = None
        vrma_raw = str(row.get("RO_Numbers", ""))
        if vrma_raw and vrma_raw not in ("nan", "None", ""):
            for ro_num in [x.strip() for x in vrma_raw.split(",")]:
                vrma = ro_vrma_lookup.get((wn, ro_num, pn_str), "")
                if vrma and vrma in haas_by_rma:
                    haas_matched = haas_by_rma[vrma][0]
                    break

        # Secondary: Part + Haas WO Id
        if haas_matched is None:
            for hid in hids:
                key = (pn_str, hid)
                if key in haas_by_part_wo:
                    haas_matched = haas_by_part_wo[key][0]
                    add_to_ro = str(haas_matched.get("RMA Order", ""))
                    break

        # Write Haas columns
        if haas_matched is not None:
            for k, idx in RMA_HAAS_COL.items():
                v = haas_matched.get(k)
                if v is not None and str(v) not in ("nan", "None", ""):
                    c = ws.cell(row=cur, column=idx, value=str(v))
                    c.fill = solid(bg); c.font = mkfont(color=DARK); c.border = THIN_BORDER

        # If secondary match, override Vendor RMA with "Add to RO: XXXXX"
        if add_to_ro:
            c = ws.cell(row=cur, column=14, value="Add to RO: " + add_to_ro)
            c.fill = solid(C_ADD_RO)
            c.font = mkfont(bold=True, color=DARK)
            c.border = THIN_BORDER

        cur += 1


# ═══════════════════════════════════════════════════════════
# SHEET: WO Gate Summary
# ═══════════════════════════════════════════════════════════
C_GATE_HDR   = "0D1F3C"   # dark navy  — WO banner
C_GATE_PASS  = "C6EFCE"   # light green
C_GATE_FAIL  = "FFD7D7"   # light red
C_GATE_FINAL_PASS = "375623"  # dark green  — WO ready
C_GATE_FINAL_FAIL = "9C0006"  # dark red    — WO blocked
C_GATE_ROW   = "F2F2F2"   # light grey  — gate rows
C_GATE_ALT   = "FFFFFF"   # white       — alternating

GS_WIDTHS = {"A":22,"B":10,"C":60}

def gs_wo_banner(ws, r, wo_num, customer, tech, subtype):
    ws.row_dimensions[r].height = 22
    label = f"{wo_num}   |   {customer}   |   {tech}   |   {subtype}"
    c = ws.cell(r, 1, label)
    c.fill = solid(C_GATE_HDR); c.font = mkfont(bold=True, color=WHITE, size=11)
    c.alignment = Alignment(vertical="center")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

def gs_gate_row(ws, r, gate_name, result, detail, is_final=False, alt=False, is_doc=False):
    ws.row_dimensions[r].height = 18 if not is_doc else 22
    bg = C_GATE_ALT if alt else C_GATE_ROW

    # Col A — Gate name
    c1 = ws.cell(r, 1, gate_name)
    c1.fill = solid(bg); c1.font = mkfont(bold=(is_final or is_doc), color=DARK)
    c1.border = THIN_BORDER; c1.alignment = Alignment(vertical="center")

    # Col B — result (colored)
    if is_final:
        res_fill = C_GATE_FINAL_PASS if "ready" in result.lower() else C_GATE_FINAL_FAIL
        res_fc   = WHITE
    elif is_doc:
        if result == "Pass":
            res_fill = C_DQ_GOOD; res_fc = DARK
        elif result == "Warn":
            res_fill = C_DQ_WARN; res_fc = DARK
        else:
            res_fill = C_DQ_BLOCK; res_fc = WHITE
    else:
        res_fill = C_GATE_PASS if result == "Pass" else (C_DQ_WARN if result == "Warn" else C_GATE_FAIL)
        res_fc   = DARK

    c2 = ws.cell(r, 2, result)
    c2.fill = solid(res_fill); c2.font = mkfont(bold=(is_final or is_doc), color=res_fc)
    c2.border = THIN_BORDER; c2.alignment = Alignment(horizontal="center", vertical="center")

    # Col C — detail
    c3 = ws.cell(r, 3, detail)
    c3.fill = solid(bg); c3.font = mkfont(color=DARK)
    c3.border = THIN_BORDER; c3.alignment = Alignment(vertical="center", wrap_text=True)

def gs_spacer(ws, r):
    ws.row_dimensions[r].height = 6
    for c in range(1, 4):
        ws.cell(r, c).fill = solid("FFFFFF")

# ── Gate logic functions ─────────────────────────────────

def gate_wo_status(wrow):
    status = str(wrow.get("WO_Status", "")).strip()
    if status == "Field Work Complete":
        return "Pass", "WO status is Field Work Complete"
    return "Fail", f"WO status is {status or 'unknown'} — not ready for approval"

def gate_destination(wrow, woli_lookup, sa_lookup, wo_customer, destin_lookup=None):
    """
    Destination fee gate — labor-based, structure-checked, same-customer dedup.
    Cross-references DESTIN-01 line item presence vs calculated fees.

    Rules:
    1. Every visit must have: outbound travel → labor → return travel.
       Missing any leg = incomplete structure flag.
    2. Qualification is based on LABOR minutes only (not travel).
       Total labor across all WOs at the same customer on the same day
       must exceed 30 minutes to justify a destination fee.
    3. If qualified, only the WO with the earliest SA dispatch that day
       carries the fee. All others must have it removed.
    4. If two WOs at the same customer each have <30 min labor but
       combined they exceed 30 min → one fee applies to the earliest.
    5. Cross-check: compare calculated fees vs DESTIN-01 line in WOLI.
       - DESTIN-01 present but shouldn't be → Fail (remove it)
       - DESTIN-01 missing but should be → Fail (add it)
       - DESTIN-01 qty doesn't match calculated fees → Fail (adjust)
       - Match → Pass
    """
    if destin_lookup is None:
        destin_lookup = {}
    wn       = str(wrow.get("WO#","")).strip().split(".")[0]
    customer = wo_customer.get(wn, "")
    wo_woli  = woli_lookup.get(wn, [])
    wo_sa    = sa_lookup.get(wn, [])
    destin_qty = destin_lookup.get(wn, 0)   # current DESTIN-01 qty on WO

    # ── No WOLI data at all ───────────────────────────────────────────────────
    if not wo_woli:
        if destin_qty > 0:
            return "Fail", f"DESTIN-01 ({destin_qty:.0f}) on WO but no WOLI lines — remove destination fee"
        return "Pass", "No WOLI lines — destination check not applicable"

    # ── Group WOLI lines by date ──────────────────────────────────────────────
    from collections import defaultdict
    days = defaultdict(lambda: {"travel_out":[], "labor":[], "travel_home":[]})
    for e in wo_woli:
        if e["date"] is None: continue
        d = e["date"]
        if e["type"] == "labor":
            days[d]["labor"].append(e)
        elif e["type"] == "travel":
            if e["travel_home"]:
                days[d]["travel_home"].append(e)
            else:
                days[d]["travel_out"].append(e)

    if not days:
        if destin_qty > 0:
            return "Fail", f"DESTIN-01 ({destin_qty:.0f}) on WO but no dated WOLI lines — remove destination fee"
        return "Pass", "No dated WOLI lines — destination check not applicable"

    # ── Check each visit date for structure and labor ─────────────────────────
    structure_issues = []
    dates_with_labor = []

    for d, legs in sorted(days.items()):
        has_out   = len(legs["travel_out"])  > 0
        has_labor = len(legs["labor"])       > 0
        has_home  = len(legs["travel_home"]) > 0
        labor_min = sum(e["duration_min"] for e in legs["labor"])

        missing = []
        if not has_out:   missing.append("outbound travel")
        if not has_labor: missing.append("labor")
        if not has_home:  missing.append("return travel")

        if missing:
            structure_issues.append(f"{d}: missing {', '.join(missing)}")

        # Flag: travel logged but no labor — likely missing WOLI entries
        if (has_out or has_home) and not has_labor:
            structure_issues.append(f"{d}: travel logged but NO labor — check for missing WOLI")

        if has_labor:
            dates_with_labor.append((d, labor_min))

    # ── No labor logged on any visit ─────────────────────────────────────────
    if not dates_with_labor:
        detail = "No labor logged — 0 destination fees apply"
        if structure_issues:
            detail = "No labor logged — " + "; ".join(structure_issues[:3])
        if destin_qty > 0:
            detail += f" | DESTIN-01 ({destin_qty:.0f}) present — REMOVE"
        return "Fail", detail

    # ── For each date with labor, do same-customer cross-WO check ────────────
    fees_this_wo = 0      # dates where this WO carries a valid fee
    remove_dates = []     # dates where fee should be removed from this WO
    no_fee_dates = []     # dates where labor doesn't qualify at all

    for visit_date, this_wo_labor in dates_with_labor:
        if not customer:
            # Can't cross-check — just check labor threshold
            if this_wo_labor > 30:
                fees_this_wo += 1
            else:
                no_fee_dates.append(visit_date)
            continue

        # Find all WOs at same customer on this date
        same_day_wos = {}
        for other_wn, other_lines in woli_lookup.items():
            other_cust = wo_customer.get(other_wn, "")
            if other_cust != customer: continue
            day_labor = sum(
                e["duration_min"] for e in other_lines
                if e["type"] == "labor" and e["date"] == visit_date
            )
            if day_labor > 0:
                same_day_wos[other_wn] = day_labor

        total_labor_day = sum(same_day_wos.values())

        if total_labor_day <= 30:
            no_fee_dates.append(visit_date)
            continue

        # Day qualifies — who carries the fee?
        candidate_sas = []
        for other_wn in same_day_wos:
            for sa in sa_lookup.get(other_wn, []):
                if sa["date"] == visit_date and sa["earliest_dt"]:
                    candidate_sas.append((sa["earliest_dt"], other_wn))

        if not candidate_sas or len(same_day_wos) == 1:
            fees_this_wo += 1
            continue

        candidate_sas.sort(key=lambda x: x[0])
        earliest_wo = candidate_sas[0][1]

        if earliest_wo == wn:
            fees_this_wo += 1
        else:
            remove_dates.append(f"{visit_date} → WO {earliest_wo}")

    # ── Build result with DESTIN-01 cross-reference ──────────────────────────
    total_visits = len(dates_with_labor)
    under_threshold = len(no_fee_dates)
    fees_apply = fees_this_wo
    removed = len(remove_dates)

    # Flag travel-only dates (zero-labor visits)
    travel_only = [si for si in structure_issues if "NO labor" in si]

    summary = f"Total visits: {total_visits}"
    if fees_apply:
        summary += f" | {fees_apply} destination fee(s) should apply"
    if under_threshold:
        summary += f" | {under_threshold} visit(s) under 30-min threshold"
    if removed:
        summary += f" | {removed} fee(s) reassigned to earlier WO"
    if travel_only:
        summary += f" | {len(travel_only)} date(s) with travel but no labor"

    # ── DESTIN-01 cross-reference ────────────────────────────────────────────
    issues = []
    if remove_dates or no_fee_dates or travel_only:
        issues.append("structure/threshold issue")

    if destin_qty > 0 and fees_apply == 0:
        summary += f" | DESTIN-01 ({destin_qty:.0f}) present but 0 fees apply — REMOVE destination fee"
        issues.append("destin-01 should not be here")
    elif destin_qty == 0 and fees_apply > 0:
        summary += f" | NO DESTIN-01 found — ADD {fees_apply} destination fee(s)"
        issues.append("destin-01 missing")
    elif destin_qty > 0 and fees_apply > 0 and int(destin_qty) != fees_apply:
        summary += f" | DESTIN-01 qty ({destin_qty:.0f}) ≠ calculated ({fees_apply}) — ADJUST"
        issues.append("destin-01 qty mismatch")
    elif destin_qty > 0 and fees_apply > 0 and int(destin_qty) == fees_apply:
        summary += f" | DESTIN-01 ({destin_qty:.0f}) matches"

    if issues:
        return "Fail", summary

    return "Pass", summary

def gate_ro_eligible_parts(parts_for_wo):
    real   = [p for p in parts_for_wo if "CORE CHARGE" not in str(p.get("PartNumber","")).upper()]
    pseudo = [p for p in parts_for_wo if "CORE CHARGE" in str(p.get("PartNumber","")).upper()]
    if not parts_for_wo:
        return "Pass", "No parts on this WO"
    if not real and pseudo:
        return "Pass", "Core charge line(s) only — no real parts"
    if not real:
        return "Pass", "No parts on this WO"
    pns = [str(p.get("PartNumber","?")) for p in real]
    return "Pass", f"{len(real)} part(s): " + ", ".join(pns)

def gate_pr_prli(parts_for_wo):
    ok = []
    missing = []          # missing PR and NOT consumed → hard fail
    van_candidates = []   # missing PR but consumed → likely van stock → warn
    for p in parts_for_wo:
        pn = str(p.get("PartNumber",""))
        if "CORE CHARGE" in pn.upper():
            continue
        pr   = val(p, "PR_Number")
        prli = val(p, "PRLI_Number")
        src  = str(p.get("PartSource",""))
        if not pr and "van" in src.lower():
            ok.append(f"{pn} (van stock)")
            continue
        if not pr or not prli:
            # Check if part is fully consumed (likely van inventory, no PR needed)
            qr = sf(p.get("QtyRequested", 0))
            qc = sf(p.get("QtyConsumed", 0))
            if qr > 0 and qc > 0 and qc >= qr:
                van_candidates.append(pn)
            else:
                missing.append(pn)
        else:
            ok.append(f"{pr} | {pn}")
    # Hard fail if any parts missing PR and NOT consumed
    if missing:
        detail = "MISSING PR: " + ", ".join(missing[:4])
        if van_candidates:
            detail += " | CONSUMED (confirm van stock): " + ", ".join(van_candidates[:4])
        if ok:
            detail += " | OK: " + ", ".join(ok[:4])
        return "Fail", detail
    # Warn if all missing-PR parts are consumed (probable van pulls)
    if van_candidates:
        detail = "No PR — consumed, confirm van stock: " + ", ".join(van_candidates[:4])
        if ok:
            detail += " | OK: " + ", ".join(ok[:4])
        return "Warn", detail
    return "Pass", ", ".join(ok[:6]) + ("..." if len(ok) > 6 else "")

def gate_consumed_vs_nnu(parts_for_wo):
    consumed = []
    nnu      = []
    partial  = []
    issues   = []
    for p in parts_for_wo:
        pn = str(p.get("PartNumber",""))
        if "CORE CHARGE" in pn.upper():
            continue
        qr = sf(p.get("QtyRequested", 0))
        qc_raw = p.get("QtyConsumed", None)
        if qc_raw is None or str(qc_raw).strip() in ("nan","None",""):
            issues.append(f"{pn} (consumed qty missing)")
            continue
        qc = sf(qc_raw)
        qn = sf(p.get("QtyNotUsed", 0))
        if qr > 0 and qc > qr:
            issues.append(f"{pn} OVER-CONSUMED ({qc:.0f} consumed vs {qr:.0f} requested)")
        elif qr > 0 and qc > 0 and qc == qr:
            consumed.append(pn)
        elif qr > 0 and qc == 0:
            nnu.append(pn)
        elif qr > 0 and qc > 0 and qc < qr:
            partial.append(f"{pn} ({qc:.0f}/{qr:.0f})")
        else:
            consumed.append(pn)
        # Qty reconciliation: consumed + not-used should equal requested
        if qr > 0 and (qc + qn) != qr and qc <= qr:
            unaccounted = qr - qc - qn
            issues.append(f"{pn} QTY MISMATCH (req {qr:.0f}, consumed {qc:.0f}, NNU {qn:.0f}, {unaccounted:.0f} unaccounted)")
    if issues:
        return "Fail", "; ".join(issues[:4])
    parts = []
    if consumed: parts.append("Consumed: " + ", ".join(consumed[:4]))
    if nnu:      parts.append("NNU: " + ", ".join(nnu[:4]))
    if partial:  parts.append("Partial: " + ", ".join(partial[:3]))
    if not parts:
        return "Pass", "No real parts to reconcile"
    return "Pass", " | ".join(parts)

# ── RO type buckets ────────────────────────────────────────────────────────
NNU_REASONS   = {"NNU", "Parts Error - NNU", "Service Error - NNU"}
CORE_REASONS  = {"50% Exchange"}
WARR_REASONS  = {"Warranty"}

# ── RO status priority (best wins) ────────────────────────────────────────
RO_STATUS_PRIORITY = {"Ready To Ship": 0, "Approved": 1, "Open": 2, "Draft": 3}

def _ro_bucket(reason):
    """Classify an RO reason into NNU / Warranty / Core bucket."""
    r = str(reason).strip()
    if r in NNU_REASONS:  return "NNU"
    if r in WARR_REASONS: return "Warranty"
    if r in CORE_REASONS: return "Core"
    return "Other"

def _best_ro_status(ro_lines):
    """Return the best status among matched ROs using priority order."""
    best = None
    best_pri = 999
    for r in ro_lines:
        st = str(r.get("Status","")).strip()
        pri = RO_STATUS_PRIORITY.get(st, 4)  # 4 = Other
        if pri < best_pri:
            best_pri = pri
            best = st
    return best or "No Match"

def gate_ro_coverage(parts_for_wo, wst, ro_df, wo_num):
    ok       = []
    issues   = []
    warnings = []
    for p in parts_for_wo:
        pn = str(p.get("PartNumber",""))
        if "CORE CHARGE" in pn.upper():
            continue
        ro_lines = get_ro(ro_df, wo_num, pn)

        # No ROs at all
        if not ro_lines:
            qc = sf(p.get("QtyConsumed", 0))
            qr = sf(p.get("QtyRequested", 0))
            is_warranty = (wst == "Warranty")
            if qr > 0 and qc == 0:
                issues.append(f"{pn}: no RO — needs NNU")
            elif is_warranty and qr > 0 and qc > 0:
                issues.append(f"{pn}: no RO — warranty consumed, needs Warranty RO")
            elif qc > 0:
                ok.append(f"{pn}: consumed, no RO needed")
            else:
                ok.append(pn)
            continue

        # Classify ROs into buckets
        buckets = set()
        bucket_counts = {}
        for r in ro_lines:
            reason = str(r.get("Reason For Return","")).strip()
            b = _ro_bucket(reason)
            buckets.add(b)
            bucket_counts[b] = bucket_counts.get(b, 0) + 1

        best_status = _best_ro_status(ro_lines)

        # Flag: multiple ROs of same type (duplicate detection — #5)
        dupe_buckets = [b for b, cnt in bucket_counts.items() if cnt > 1]
        if dupe_buckets:
            for db in dupe_buckets:
                warnings.append(f"{pn}: {bucket_counts[db]} duplicate {db} ROs")

        # Flag: conflicting reason buckets
        real_buckets = buckets - {"Other"}
        if len(real_buckets) > 1:
            issues.append(f"{pn}: RO conflict — {', '.join(sorted(real_buckets))} on same part")
            continue

        # Single bucket — validate it matches the scenario
        bucket = real_buckets.pop() if real_buckets else "Other"
        qr = sf(p.get("QtyRequested", 0))
        qc = sf(p.get("QtyConsumed", 0))
        is_consumed = (qr > 0 and qc > 0 and qc >= qr)
        is_nnu = (qr > 0 and qc == 0)
        is_warranty = (wst == "Warranty")

        if is_nnu and bucket == "NNU":
            ok.append(f"{pn}: NNU RO OK ({best_status})")
        elif is_nnu and bucket != "NNU":
            issues.append(f"{pn}: not consumed but has {bucket} RO — needs NNU instead")
        elif is_consumed and is_warranty and bucket == "Warranty":
            ok.append(f"{pn}: Warranty RO OK ({best_status})")
        elif is_consumed and bucket == "Core":
            ok.append(f"{pn}: Core RO OK ({best_status})")
        elif is_consumed and bucket == "NNU":
            warnings.append(f"{pn}: consumed but has NNU RO — remove NNU")
        elif is_consumed and is_warranty and bucket != "Warranty":
            issues.append(f"{pn}: warranty consumed but has {bucket} RO — needs Warranty RO")
        elif is_consumed:
            ok.append(f"{pn}: consumed + {bucket} RO ({best_status})")
        else:
            ok.append(f"{pn}: {bucket} ({best_status})")

    if issues:
        detail = "NEED ACTION: " + "; ".join(issues[:4])
        if warnings:
            detail += " | WARN: " + "; ".join(warnings[:2])
        if ok:
            detail += " | OK: " + ", ".join(ok[:3])
        return "Fail", detail
    if warnings:
        detail = "; ".join(warnings[:4])
        if ok:
            detail += " | OK: " + ", ".join(ok[:3])
        return "Warn", detail
    if not ok:
        return "Pass", "No parts requiring RO coverage"
    return "Pass", ", ".join(ok[:5]) + ("..." if len(ok) > 5 else "")

def gate_ro_status(parts_for_wo, ro_df, wo_num):
    good = []
    bad  = []
    for p in parts_for_wo:
        pn = str(p.get("PartNumber",""))
        if "CORE CHARGE" in pn.upper():
            continue
        ro_lines = get_ro(ro_df, wo_num, pn)
        for r in ro_lines:
            st = str(r.get("Status","")).strip()
            ro_num = str(r.get("Return Order Number","")).strip()
            if st and st not in ("nan","None",""):
                if st in GOOD_STAT:
                    good.append(f"{pn} \u2192 {ro_num} ({st})")
                else:
                    bad.append(f"{pn} \u2192 {ro_num} ({st})")
    if bad:
        detail = "FIX: " + "; ".join(bad[:4])
        if good:
            detail += " | OK: " + ", ".join(good[:3])
        return "Fail", detail
    if not good:
        return "Pass", "No ROs to validate"
    return "Pass", ", ".join(good[:5]) + ("..." if len(good) > 5 else "")

# ── Gate 8: Copy/Paste CA Detection ──────────────────────
def _build_tech_ca_index(wo_df, ca_lk):
    """Build tech -> [(wo#, ca_text)] index for duplicate CA detection."""
    tech_cas = {}
    for _, r in wo_df.iterrows():
        wn   = str(r.get("WO#", "")).strip().split(".")[0]
        tech = str(r.get("WO_Technician", "")).strip()
        if not tech or tech in ("nan", "None", ""):
            continue
        _, ca = get_ca(ca_lk, wn)
        ca_norm = re.sub(r'\s+', ' ', ca.lower().strip())
        if len(ca_norm) < 20:
            continue
        tech_cas.setdefault(tech, []).append((wn, ca_norm))
    return tech_cas

def gate_ca_duplicate(wo_num, tech, tech_ca_index):
    """Flag when a tech uses identical CA text across multiple WOs."""
    tech = str(tech).strip()
    if not tech or tech in ("nan", "None", ""):
        return "Pass", "No technician assigned"
    entries = tech_ca_index.get(tech, [])
    if len(entries) < 2:
        return "Pass", "Only one WO for this tech — no comparison needed"
    wn = str(wo_num).strip().split(".")[0]
    this_ca = None
    for w, ca in entries:
        if w == wn:
            this_ca = ca
            break
    if this_ca is None or len(this_ca) < 20:
        return "Pass", "CA text too short to compare"
    dupes = [w for w, ca in entries if ca == this_ca and w != wn]
    if dupes:
        wo_list = ", ".join(dupes[:5])
        extra = f" +{len(dupes)-5} more" if len(dupes) > 5 else ""
        return "Warn", f"Identical CA text found on {tech}'s WO(s): {wo_list}{extra} — verify not copy/pasted"
    return "Pass", "CA text is unique for this tech"

# ── Main builder ─────────────────────────────────────────

def build_gate_summary(ws, data):
    wo_all = data["WO_Output"]
    wo_df  = wo_all[wo_all["WO_Status"].isin(AUDIT_STATUSES)] if "WO_Status" in wo_all.columns else wo_all
    pa_df  = data["Parts_Output"]
    ro_df  = data["RO"]
    ca_lk  = data.get("ca_lookup", {})

    # Build destination lookups once for all WOs
    woli_lk, sa_lk, wo_cust, destin_lk = build_destination_lookups(data)

    # Build tech CA index once for copy/paste detection
    tech_ca_idx = _build_tech_ca_index(wo_df, ca_lk)

    # Header row
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26
    setw(ws, GS_WIDTHS)
    for ci, label in enumerate(["Gate", "Result", "Detail"], 1):
        c = ws.cell(1, ci, label)
        c.fill = solid(C_GATE_HDR); c.font = mkfont(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER

    cur = 2
    for _, wrow in wo_df.iterrows():
        wn   = str(wrow.get("WO#",""))
        cust = str(wrow.get("Customer",""))
        tech = str(wrow.get("WO_Technician",""))
        sub  = str(wrow.get("WO_Subtype",""))

        # Get parts for this WO as list of dicts
        wp = pa_df[pa_df["WO#"] == wn] if len(pa_df) else pd.DataFrame()
        parts = wp.to_dict("records")

        # Get CA text for this WO
        cause, ca = get_ca(ca_lk, wn)

        # WO banner
        gs_wo_banner(ws, cur, wn, cust, tech, sub)
        cur += 1

        # Run standard gates
        gates = [
            ("WO Status Ready",              gate_wo_status(wrow.to_dict())),
            ("Destination Review",           gate_destination(wrow.to_dict(), woli_lk, sa_lk, wo_cust, destin_lk)),
            ("RO-Eligible Parts Identified", gate_ro_eligible_parts(parts)),
            ("PR / PRLI Integrity",          gate_pr_prli(parts)),
            ("Consumed vs Not Used",         gate_consumed_vs_nnu(parts)),
            ("Required RO Coverage",         gate_ro_coverage(parts, sub, ro_df, wn)),
            ("RO Status Valid",              gate_ro_status(parts, ro_df, wn)),
        ]

        failed_gates = [g for g, (r, _) in gates if r == "Fail"]

        for i, (gate_name, (result, detail)) in enumerate(gates):
            gs_gate_row(ws, cur, gate_name, result, detail, alt=(i % 2 == 0))
            cur += 1

        # ── Documentation Quality Gate (second-to-last, before Final) ─────
        # Check if WO has any parts sold (for part# warning)
        wo_has_parts = len(parts) > 0
        dq_result, dq_detail, dq_scores = dq_gate(cause, ca, sub, wn, has_parts=wo_has_parts)
        is_doc_skip = dq_scores.get("skip", False)

        # Merge CA duplicate check into doc quality as a warning
        ca_dup_result, ca_dup_detail = gate_ca_duplicate(wn, tech, tech_ca_idx)
        if ca_dup_result == "Warn":
            dq_detail += f" | COPY/PASTE WARNING: {ca_dup_detail}"
            if dq_result == "Pass":
                dq_result = "Warn"

        # Warn counts as pass for Final gate — only Fail blocks
        if dq_result == "Fail":
            failed_gates.append("Documentation Quality")

        gs_gate_row(ws, cur, "Documentation Quality", dq_result, dq_detail,
                    is_doc=(not is_doc_skip), alt=False)
        cur += 1

        # ── Final row ──────────────────────────────────────────────────────
        if failed_gates:
            final_result = "WO blocked"
            final_detail = "Blocked by: " + ", ".join(failed_gates)
        else:
            final_result = "WO ready to send to D365"
            final_detail = "All gates passed"

        gs_gate_row(ws, cur, "Final", final_result, final_detail, is_final=True)
        cur += 1

        # Spacer
        gs_spacer(ws, cur)
        cur += 1

    ws.sheet_properties.tabColor = "A9D08E"

def build_doc_quality_sheet(ws, data):
    """
    Doc Quality sheet — simplified dual-rubric display.
    Shows score, grade, gate result, elements, signals, flags per WO.
    """
    wo_all = data["WO_Output"]
    wo_df  = wo_all[wo_all["WO_Status"].isin(AUDIT_STATUSES)] if "WO_Status" in wo_all.columns else wo_all
    ca_lk  = data.get("ca_lookup", {})
    pa_df  = data.get("Parts_Output", pd.DataFrame())

    from openpyxl.utils import get_column_letter

    COLS = ["WO#", "Technician", "Subtype", "Gate Result", "Grade", "Score /100",
            "Type", "Arrival/Finding", "Work Performed", "Closure",
            "Observed (W)", "Tested (W)", "Bonus Signals",
            "Warnings", "Auto-Fail Flags"]
    WIDTHS = [10, 24, 10, 12, 8, 12, 10, 13, 14, 12, 13, 12, 28, 30, 38]

    for i, w in enumerate(WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Title
    ws.merge_cells("A1:O1")
    t = ws.cell(1, 1, "Documentation Quality Gate — Scoring Breakdown")
    t.fill = solid("0D1F3C"); t.font = mkfont(bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:O2")
    sub_cell = ws.cell(2, 1,
        f"Block: any required element missing or score < {DQ_BLOCK_THRESHOLD}  |  "
        f"Warn: score < {DQ_WARN_THRESHOLD}  |  "
        "Paid/Courtesy: Arrival(25) + Work(35) + Closure(30) + Bonus(10)  |  "
        "Warranty: Observed(20) + Tested(25) + Work(25) + Closure(20) + Bonus(10)")
    sub_cell.fill = solid("1F4E79")
    sub_cell.font = Font(italic=True, color="DDDDDD", size=9, name="Arial")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 26

    for c, label in enumerate(COLS, 1):
        cell = ws.cell(3, c, label)
        cell.fill = solid("365F91"); cell.font = mkfont(bold=True, color=WHITE, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = "A4"

    def grade_fill(g):
        return {"A":"375623","B":"C6EFCE","C":"FFEB9C","D":"F4B942","F":"FFC7CE"}.get(g,"F2F2F2")
    def grade_fc(g):
        return WHITE if g in ("A","F") else DARK
    def result_fill(r):
        return {"Pass":"C6EFCE","Warn":"FFD966","Fail":"FFC7CE","N/A":"D0D0D0"}.get(r,"F2F2F2")
    def yn_fill(v):
        return "C6EFCE" if v else "FFC7CE"
    def yn_txt(v):
        return "Yes" if v else "No"

    row = 4
    passed = warned = blocked = skipped = 0

    for _, wrow in wo_df.iterrows():
        wn       = str(wrow.get("WO#", ""))
        tech     = str(wrow.get("WO_Technician", ""))
        sub_type = str(wrow.get("WO_Subtype", ""))

        cause, ca = get_ca(ca_lk, wn)
        wp = pa_df[pa_df["WO#"] == wn] if len(pa_df) else pd.DataFrame()
        has_parts = len(wp) > 0

        dq_result, dq_detail, sc = dq_gate(cause, ca, sub_type, wn, has_parts=has_parts)
        row_bg = "F2F2F2" if row % 2 == 0 else "FFFFFF"

        def wc(col, v, bg=None, bold=False, align="center", fc=DARK, wrap=False):
            cell = ws.cell(row, col, v)
            cell.fill   = solid(bg or row_bg)
            cell.font   = mkfont(bold=bold, color=fc, size=9)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)

        if sc.get("skip"):
            skipped += 1
            wc(1, wn); wc(2, tech, align="left"); wc(3, sub_type)
            wc(4, "N/A", bg="D0D0D0"); wc(5, "—", bg="D0D0D0"); wc(6, "—", bg="D0D0D0")
            for col in range(7, 16): wc(col, "—", bg="D0D0D0")
        else:
            s100  = sc.get("score_100", 0)
            grade = sc.get("grade", "F")
            af    = sc.get("auto_fails", [])
            warns = sc.get("warnings", [])
            sigs  = sc.get("signals", [])
            elems = sc.get("elements", {})
            is_w  = sc.get("is_warranty", False)
            rtype = "Warranty" if is_w else "Paid/Courtesy"

            if dq_result == "Fail":    blocked += 1
            elif dq_result == "Warn":  warned  += 1
            else:                      passed  += 1

            wc(1,  wn)
            wc(2,  tech, align="left")
            wc(3,  sub_type)
            wc(4,  dq_result, bg=result_fill(dq_result), bold=True,
               fc="9C0006" if dq_result=="Fail" else DARK)
            wc(5,  grade,  bg=grade_fill(grade), bold=True, fc=grade_fc(grade))
            wc(6,  s100,   bg=grade_fill(grade), bold=True)
            wc(7,  rtype)

            # Element scores
            if is_w:
                wc(8,  f"{elems.get('observed',0)}/20", bg=yn_fill(elems.get('has_observed',False)))
                wc(9,  f"{elems.get('tested',0)}/25",   bg=yn_fill(elems.get('has_tested',False)))
                wc(10, f"{elems.get('work',0)}/25",     bg=yn_fill(elems.get('has_work',False)))
                wc(11, f"{elems.get('closure',0)}/20",  bg=yn_fill(elems.get('has_closure',False)))
                wc(12, "—", bg=row_bg)  # Observed col
                # Swap cols 11/12 — for warranty: 11=work already, 12=closure not needed
            else:
                wc(8,  f"{elems.get('arrival',0)}/25",  bg=yn_fill(elems.get('has_arrival',False)))
                wc(9,  f"{elems.get('work',0)}/35",     bg=yn_fill(elems.get('has_work',False)))
                wc(10, f"{elems.get('closure',0)}/30",  bg=yn_fill(elems.get('has_closure',False)))
                wc(11, "N/A", bg="D0D0D0")   # Observed — paid only
                wc(12, "N/A", bg="D0D0D0")   # Tested  — paid only

            wc(13, ", ".join(sigs) if sigs else "—", align="left", wrap=True)
            wc(14, " | ".join(warns) if warns else "—",
               bg="FFD966" if warns else row_bg, align="left", wrap=True)
            wc(15, " | ".join(af) if af else "—",
               bg="FFC7CE" if af else row_bg, align="left", wrap=True)

        ws.row_dimensions[row].height = 16
        row += 1

    # Summary footer
    row += 1
    ws.merge_cells(f"A{row}:G{row}")
    total = passed + warned + blocked
    txt   = (f"Scored: {total}  |  Passed: {passed}  |  Warned: {warned}  |  "
             f"Blocked: {blocked}  |  Excluded (N/A): {skipped}")
    sc_c = ws.cell(row, 1, txt)
    sc_c.fill = solid("1F4E79"); sc_c.font = mkfont(bold=True, color=WHITE, size=10)
    sc_c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    ws.sheet_properties.tabColor = "2E75B6"


INPUT_DIR  = "Input"
OUTPUT_DIR = "Output"

def main():
    # ── Create folders ──────────────────────────────────────────────────
    for folder in (INPUT_DIR, OUTPUT_DIR):
        os.makedirs(folder, exist_ok=True)

    # ── Find the WO report in Input/ ────────────────────────────────────
    xlsx_files = [f for f in os.listdir(INPUT_DIR)
                  if f.lower().endswith(".xlsx") and not f.startswith("~")]

    # Separate Haas RMA file from WO report
    haas_candidates = [f for f in xlsx_files if "rma" in f.lower()]
    wo_candidates   = [f for f in xlsx_files if f not in haas_candidates]

    if not wo_candidates:
        print("ERROR: No .xlsx file found in '%s/' folder." % INPUT_DIR)
        print("       Drop your WO Approval Report into the Input folder and re-run.")
        return
    if len(wo_candidates) > 1:
        print("WARNING: Multiple .xlsx files in Input/ — using the first one:")
        for x in wo_candidates:
            print("  • %s" % x)

    f  = os.path.join(INPUT_DIR, wo_candidates[0])
    hf = os.path.join(INPUT_DIR, haas_candidates[0]) if haas_candidates else None

    print("Reading: %s" % f)
    d = load_data(f)

    # ── Validate report format ─────────────────────────────────────────
    # Required sheets: WO_Output and Parts_Output must have rows and key columns
    validation_errors = []
    wo_out = d.get("WO_Output", pd.DataFrame())
    pa_out = d.get("Parts_Output", pd.DataFrame())
    if len(wo_out) == 0:
        validation_errors.append("WO_Output sheet is missing or empty — this file may not be a WO Approval Report.")
    elif "WO#" not in wo_out.columns:
        validation_errors.append("WO_Output sheet has no 'WO#' column — expected a WO Approval Report format.")
    if len(pa_out) == 0:
        validation_errors.append("Parts_Output sheet is missing or empty.")
    elif "WO#" not in pa_out.columns:
        validation_errors.append("Parts_Output sheet has no 'WO#' column.")
    if validation_errors:
        print("\nERROR: File does not appear to be a valid WO Approval Report:")
        for e in validation_errors:
            print("  • %s" % e)
        print("\nExpected sheets: WO_Output, Parts_Output, RO, WO, WOLI, SA")
        print("Drop the correct Salesforce export into the Input/ folder and re-run.")
        return

    # Report audit scope
    if "WO_Status" in wo_out.columns:
        audit_count = len(wo_out[wo_out["WO_Status"].isin(AUDIT_STATUSES)])
        other_count = len(wo_out) - audit_count
        print(f"  Audit scope       : {audit_count:6d} WOs at {', '.join(AUDIT_STATUSES)}")
        if other_count:
            print(f"  Cross-ref only    : {other_count:6d} WOs (other statuses — used for destination checks)")

    d["Parts_Output"] = enrich_desc(d["Parts_Output"], d.get("PARTS_SOLD"))
    d["Parts_Output"] = enrich_rma(d["Parts_Output"], d["RO"])

    # Build CA lookup from WO sheet (Cause + Corrective Action)
    d["ca_lookup"] = build_ca_lookup(d)
    ca_count = sum(1 for c, ca in d["ca_lookup"].values() if ca.strip())
    print("  CA lookup         : %6d WOs with Corrective Action" % ca_count)

    haas_df = None
    if hf:
        haas_df = pd.read_excel(hf, header=1, dtype=str)
        print("  Haas RMA file     : %6d rows  (%s)" % (len(haas_df), os.path.basename(hf)))
    else:
        print("  Haas RMA file     : not found in Input/ — skipping merge")

    print("\nBuilding workbook...")
    wb = Workbook(); wb.remove(wb.active)
    built  = []
    failed = []

    def try_sheet(name, build_func, *args, **kwargs):
        """Build a sheet; on failure log the error and remove the empty tab."""
        try:
            ws = wb.create_sheet(name)
            build_func(ws, *args, **kwargs)
            built.append(name)
            print("  ✓ %s" % name)
        except Exception as e:
            failed.append((name, str(e)))
            if name in wb.sheetnames:
                del wb[name]
            print("  ✗ %s — ERROR: %s" % (name, e))

    try_sheet("WO Approval View", build_approval, d)
    try_sheet("WO Summary",       build_summary, d["WO_Output"])
    try_sheet("Parts Detail",     build_parts, d["Parts_Output"], d["WO_Output"], d["RO"], data=d)
    if haas_df is not None:
        try_sheet("Parts + Haas RMA", build_rma_merged, d, haas_df)
    try_sheet("WO Gate Summary",  build_gate_summary, d)
    try_sheet("Doc Quality",      build_doc_quality_sheet, d)
    try_sheet("Key",              build_key, has_haas=(haas_df is not None))

    # Move WO Gate Summary to first position if it built
    if "WO Gate Summary" in wb.sheetnames:
        idx = wb.sheetnames.index("WO Gate Summary")
        wb.move_sheet("WO Gate Summary", offset=-idx)

    # ── Always save whatever we built ───────────────────────────────────
    # Name output after the input file for traceability
    base = os.path.splitext(wo_candidates[0])[0]
    out = os.path.join(OUTPUT_DIR, base + "_AUDIT.xlsx")
    if built:
        wb.save(out)
        print("\nSaved: %s" % out)
        print("  Sheets built: %d/%d — %s" % (
            len(built), len(built) + len(failed), ", ".join(built)))
    else:
        print("\nERROR: No sheets could be built — file not saved.")

    if failed:
        print("\n  ⚠ FAILED sheets:")
        for name, err in failed:
            print("    ✗ %s — %s" % (name, err))

if __name__ == "__main__":
    main()
