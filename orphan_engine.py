"""
Orphan Work Order Analyzer Engine v1.0
Evaluates WO/SA data against HEALTHY/ORPHAN rule chain.
Produces a formatted XLSX report matching the reference output.
"""
import io
import json
from datetime import datetime, date
from collections import defaultdict, Counter

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pytz

# ── Constants ─────────────────────────────────────────────────────────────
TZ = pytz.timezone("America/New_York")

EXCLUDE_WO_STATUSES = {
    "Canceled", "Cancelled", "Closed",
    "Closed - Canceled", "Closed - Cancelled",
}
CLOSED_SA_STATUSES = {
    "Canceled", "Cancelled", "Closed",
    "Closed - Canceled", "Closed - Cancelled",
}

# Expected input columns (by position)
# Column 5 (SA Status) may appear as "Status", "Status.1", or "Status2"
# depending on the Salesforce export method. All are accepted.
EXPECTED_HEADERS = [
    "Work Order Number",
    "Account: Account Name",
    "Status",           # WO Status
    "Appointment Number",
    None,               # SA Status — validated separately below
    "Earliest Start Permitted",
    "Scheduled Start",
    "Due Date",
]
SA_STATUS_ACCEPTED = {"status", "status.1", "status2"}


# ── Helpers ───────────────────────────────────────────────────────────────
def _parse_datetime(val):
    """Parse a date/datetime value into a Python datetime, or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime(val.year, val.month, val.day)
    v = str(val).strip()
    if v == "" or v.lower() == "none":
        return None
    for fmt in (
        "%m/%d/%Y %I:%M %p",
        "%m/%d/%Y %H:%M",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
    ):
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def _to_date(val):
    """Parse to date only (no time component)."""
    dt = _parse_datetime(val)
    return dt.date() if dt else None


def _fmt_datetime(val):
    """Format a datetime for the output report (YYYY-MM-DD HH:MM) or None."""
    dt = _parse_datetime(val)
    if dt is None:
        return None
    return dt.strftime("%Y-%m-%d %H:%M")


def _clean(val):
    """Strip whitespace from a string value."""
    if val is None:
        return ""
    return str(val).strip()


# ── Core engine ───────────────────────────────────────────────────────────
def run_orphan_analysis(wo_file, source_filename=None):
    """
    Run the orphan work order analysis.

    Parameters
    ----------
    wo_file : file-like or str
        An .xlsx file (path or file-like object from Streamlit uploader).
    source_filename : str, optional
        Name to record in the Summary sheet. Defaults to "uploaded file".

    Returns
    -------
    dict with keys:
        validation_ok    : bool
        validation_errors: list[str]  (only if validation_ok is False)
        total_wos        : int
        orphan_count     : int
        healthy_count    : int
        orphan_rate      : float
        reason_counts    : dict[str, int]
        orphans          : list[dict]   (each orphan row)
        xlsx_bytes       : bytes        (the formatted report)
        log              : list[str]
    """
    today = datetime.now(TZ).date()
    log = []
    log.append(f"[INFO] Analysis date (ET): {today}")

    if source_filename is None:
        source_filename = getattr(wo_file, "name", "uploaded file")

    # ── Load workbook ─────────────────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(wo_file, data_only=True)
    except Exception as e:
        return {
            "validation_ok": False,
            "validation_errors": [f"Cannot open file: {e}"],
        }

    ws = wb.active
    if ws is None or ws.max_row is None or ws.max_row < 2:
        return {
            "validation_ok": False,
            "validation_errors": ["File has no data rows."],
        }

    # ── Validate headers ──────────────────────────────────────────────
    header_row = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    errors = []
    for i, expected in enumerate(EXPECTED_HEADERS):
        got = _clean(header_row[i]).lower() if i < len(header_row) else ""
        if expected is None:
            # SA Status column — accept any of the known variants
            if got not in SA_STATUS_ACCEPTED:
                errors.append(f"Column {i+1}: expected 'Status' or 'Status2', got '{header_row[i] if i < len(header_row) else '(missing)'}'")
        elif got != expected.lower():
            errors.append(f"Column {i+1}: expected '{expected}', got '{header_row[i] if i < len(header_row) else '(missing)'}'")

    if errors:
        return {"validation_ok": False, "validation_errors": errors}

    log.append(f"[OK]  Headers validated ({len(header_row)} columns)")

    # ── Parse rows into WO groups ─────────────────────────────────────
    wos = {}
    row_count = 0
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        wo_num   = _clean(row[0])
        acct     = _clean(row[1])
        wo_stat  = _clean(row[2])
        appt_num = _clean(row[3])
        sa_stat  = _clean(row[4])
        esp_raw  = row[5]
        sched_raw = row[6]
        due_raw  = row[7]

        if not wo_num:
            continue
        row_count += 1

        # Treat blank / "None" SA status as New
        if sa_stat == "" or sa_stat.lower() == "none":
            sa_stat = "New"

        sched_dt = _parse_datetime(sched_raw)
        due_dt   = _parse_datetime(due_raw)

        if wo_num not in wos:
            wos[wo_num] = {
                "acct": acct,
                "wo_status": wo_stat,
                "sas": [],
            }

        wos[wo_num]["sas"].append({
            "appt_num":    appt_num,
            "sa_status":   sa_stat,
            "sched_start": sched_dt,
            "due_date":    due_dt,
        })

    log.append(f"[OK]  Parsed {row_count} data rows -> {len(wos)} unique WOs")

    # ── Phase 1: exclude closed/canceled WOs ──────────────────────────
    excluded = {k for k, v in wos.items() if v["wo_status"] in EXCLUDE_WO_STATUSES}
    active_wos = {k: v for k, v in wos.items() if k not in excluded}
    log.append(f"[OK]  Excluded {len(excluded)} Canceled/Closed WOs")
    log.append(f"[OK]  {len(active_wos)} active WOs to evaluate")

    # ── Helpers for evaluation ────────────────────────────────────────
    def get_latest_sa(sas):
        """Latest SA by Scheduled Start desc, then Appointment Number desc."""
        def sort_key(sa):
            sd = sa["sched_start"] if sa["sched_start"] else datetime.min
            an = sa["appt_num"] if sa["appt_num"] else ""
            return (sd, an)
        return sorted(sas, key=sort_key, reverse=True)[0]

    def get_wo_due_date(sas):
        """Get the latest due date across all SAs for this WO."""
        dates = [sa["due_date"] for sa in sas if sa["due_date"] is not None]
        return max(dates) if dates else None

    def has_no_sa(sa):
        return sa["appt_num"] == ""

    def has_future_or_active_sa(sas):
        for sa in sas:
            st = sa["sa_status"]
            sd = sa["sched_start"].date() if sa["sched_start"] else None
            dd = sa["due_date"].date() if sa["due_date"] else None
            if st == "New" and dd and dd >= today:
                return True
            if st in ("Scheduled", "Dispatched") and sd and sd >= today:
                return True
            if st in ("Onsite", "Travel", "Cannot Complete") and sd and sd == today:
                return True
        return False

    # ── Phase 2: evaluate each WO ─────────────────────────────────────
    orphans = []
    healthy_count = 0
    reason_counts = Counter()

    for wo_num, wo_data in active_wos.items():
        sas = wo_data["sas"]
        latest_sa = get_latest_sa(sas)
        sa_status = latest_sa["sa_status"]
        sched_start = latest_sa["sched_start"]
        due_date = get_wo_due_date(sas)

        sched_date = sched_start.date() if sched_start else None
        due_date_d = due_date.date() if due_date else None

        result = None
        reason = None

        # ── HEALTHY checks ────────────────────────────────────────────
        if sa_status == "New" and due_date_d and due_date_d >= today:
            result = "HEALTHY"
        elif sa_status in ("Scheduled", "Dispatched") and sched_date and sched_date >= today:
            result = "HEALTHY"
        elif sa_status in ("Onsite", "Travel", "Cannot Complete") and sched_date and sched_date == today:
            result = "HEALTHY"

        # ── ORPHAN checks (precedence) ────────────────────────────────
        if result is None:
            if has_no_sa(latest_sa) and (due_date_d is None or due_date_d < today):
                result = "ORPHAN"
                reason = "No Service Appointment and no future Due Date"
            elif sa_status in ("Scheduled", "Dispatched") and sched_date and sched_date < today:
                result = "ORPHAN"
                reason = "Scheduled or Dispatched Service Appointment is in the past"
            elif all(sa["sa_status"] in CLOSED_SA_STATUSES for sa in sas):
                result = "ORPHAN"
                reason = "All Service Appointments are canceled or closed"
            elif any(sa["sa_status"] == "Completed" for sa in sas) and not has_future_or_active_sa(sas):
                result = "ORPHAN"
                reason = "Completed Service Appointment with no follow-up scheduled"
            elif not has_future_or_active_sa(sas):
                result = "ORPHAN"
                reason = "No qualifying future or active Service Appointment"
            else:
                result = "HEALTHY"

        if result == "HEALTHY":
            healthy_count += 1
        else:
            reason_counts[reason] += 1
            orphans.append({
                "wo_num":      wo_num,
                "wo_status":   wo_data["wo_status"],
                "sa_status":   sa_status,
                "sched_start": sched_start,
                "due_date":    due_date,
                "reason":      reason,
            })

    # Sort orphans by WO number
    orphans.sort(key=lambda x: x["wo_num"])

    total_active = len(active_wos)
    orphan_count = len(orphans)
    orphan_rate = orphan_count / total_active if total_active > 0 else 0.0

    log.append(f"[OK]  Evaluation complete: {healthy_count} HEALTHY, {orphan_count} ORPHAN")
    for reason, cnt in reason_counts.most_common():
        log.append(f"       [{cnt}] {reason}")

    # ── Build XLSX report ─────────────────────────────────────────────
    out_wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(patternType="solid", fgColor="1F4E78")
    header_align = Alignment(horizontal="center")
    date_align = Alignment(horizontal="left")

    # ── Summary sheet ─────────────────────────────────────────────────
    ws_sum = out_wb.active
    ws_sum.title = "Summary"

    summary_data = [
        ("Metric", "Value"),
        ("Total Work Orders (after WO status filtering)", total_active),
        ("Orphaned Work Orders", orphan_count),
        ("Orphan Rate %", orphan_rate),
        ("Generated From", source_filename),
    ]
    for r_idx, (metric, value) in enumerate(summary_data, start=1):
        ws_sum.cell(row=r_idx, column=1, value=metric)
        ws_sum.cell(row=r_idx, column=2, value=value)

    # Header row formatting
    for col in (1, 2):
        cell = ws_sum.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Orphan Rate % formatting
    ws_sum.cell(row=4, column=2).number_format = "0.0%"

    # Column widths
    ws_sum.column_dimensions["A"].width = 42.0
    ws_sum.column_dimensions["B"].width = 24.0

    # ── Orphans sheet ─────────────────────────────────────────────────
    ws_orph = out_wb.create_sheet("Orphans")

    orphan_headers = [
        "Work Order Number",
        "WO Status",
        "SA Status (latest)",
        "Scheduled Start (latest)",
        "Due Date",
        "Orphan Reason",
    ]
    col_widths = {
        "A": 18.0, "B": 16.0, "C": 20.0,
        "D": 24.0, "E": 20.0, "F": 52.0,
    }

    # Write headers
    for c_idx, hdr in enumerate(orphan_headers, start=1):
        cell = ws_orph.cell(row=1, column=c_idx, value=hdr)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    # Write orphan rows
    for r_idx, orph in enumerate(orphans, start=2):
        ws_orph.cell(row=r_idx, column=1, value=orph["wo_num"])
        ws_orph.cell(row=r_idx, column=2, value=orph["wo_status"])
        ws_orph.cell(row=r_idx, column=3, value=orph["sa_status"])

        # Scheduled Start — formatted as string "YYYY-MM-DD HH:MM" or None
        sched_val = _fmt_datetime(orph["sched_start"])
        cell_d = ws_orph.cell(row=r_idx, column=4, value=sched_val)
        cell_d.alignment = date_align

        # Due Date — formatted as string "YYYY-MM-DD HH:MM" or None
        due_val = _fmt_datetime(orph["due_date"])
        cell_e = ws_orph.cell(row=r_idx, column=5, value=due_val)
        cell_e.alignment = date_align

        ws_orph.cell(row=r_idx, column=6, value=orph["reason"])

    # Column widths
    for col_letter, width in col_widths.items():
        ws_orph.column_dimensions[col_letter].width = width

    # ── Serialize to bytes ────────────────────────────────────────────
    buf = io.BytesIO()
    out_wb.save(buf)
    xlsx_bytes = buf.getvalue()

    log.append(f"[OK]  XLSX report generated ({len(xlsx_bytes):,} bytes)")

    return {
        "validation_ok": True,
        "total_wos":     total_active,
        "orphan_count":  orphan_count,
        "healthy_count": healthy_count,
        "orphan_rate":   orphan_rate,
        "reason_counts": dict(reason_counts),
        "orphans":       orphans,
        "xlsx_bytes":    xlsx_bytes,
        "log":           log,
    }


# ── CLI test harness ──────────────────────────────────────────────────────
if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else None
    if path is None:
        print("Usage: python orphan_engine.py <input.xlsx>")
        sys.exit(1)

    result = run_orphan_analysis(path, source_filename=path.split("/")[-1])
    if not result.get("validation_ok"):
        print("VALIDATION FAILED:")
        for e in result["validation_errors"]:
            print(f"  {e}")
        sys.exit(1)

    for line in result["log"]:
        print(line)

    # Write output
    out_path = path.replace(".xlsx", "") + "_ORPHAN_REPORT.xlsx"
    with open(out_path, "wb") as f:
        f.write(result["xlsx_bytes"])
    print(f"\nReport written to: {out_path}")
